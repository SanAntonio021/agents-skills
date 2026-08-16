from __future__ import annotations

import errno
import hashlib
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch


SKILL_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = SKILL_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from libreoffice_headless import convert, find_soffice  # noqa: E402
from merge_formula_caches import merge_caches  # noqa: E402
from ooxml_common import (  # noqa: E402
    MAIN_NS,
    has_formula_cache,
    load_package,
    parse_xml,
    qn,
    resolve_sheet_parts,
    serialize_xml,
    worksheet_cells,
    write_package,
)
from patch_ooxml import patch_workbook  # noqa: E402
import publish_output as publisher  # noqa: E402
from verify_pdf import find_pdftoppm, inspect_pdf  # noqa: E402
from verify_xlsx import compare_workbooks, inspect_workbook  # noqa: E402


class WorkbookToolTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory(prefix="xlsx-skill-test-")
        self.root = Path(self.temp_dir.name)
        self.source = self.root / "source.xlsx"
        self._build_workbook(self.source)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    @staticmethod
    def _build_workbook(path: Path) -> None:
        from openpyxl import Workbook
        from openpyxl.chart import BarChart, Reference
        from openpyxl.comments import Comment
        from openpyxl.worksheet.datavalidation import DataValidation

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Template"
        sheet["A1"] = "Audit workbook"
        sheet.merge_cells("A1:E1")
        sheet["B2"] = "Old value"
        sheet["C2"] = 1
        sheet["D2"] = 2
        sheet["E2"] = "=SUM(C2:D2)"
        sheet["B2"].comment = Comment("Keep this comment", "Tester")
        validation = DataValidation(type="list", formula1='"Domestic,Imported"')
        sheet.add_data_validation(validation)
        validation.add("F2:F3")
        sheet.row_dimensions[2].height = 20
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_area = "A1:J5"
        sheet.print_title_rows = "1:1"
        chart = BarChart()
        chart.add_data(Reference(sheet, min_col=3, max_col=4, min_row=2, max_row=2))
        sheet.add_chart(chart, "H2")
        workbook.save(path)

    @staticmethod
    def _build_filter_workbook(
        path: Path, *, worksheet_filter_ref: str | None, table_ref: str | None
    ) -> None:
        from openpyxl import Workbook
        from openpyxl.worksheet.table import Table

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Filters"
        for row in range(1, 11):
            for column in range(1, 9):
                sheet.cell(row=row, column=column, value=f"R{row}C{column}")
        if table_ref is not None:
            sheet.add_table(Table(displayName="FilterTable", ref=table_ref))
        if worksheet_filter_ref is not None:
            sheet.auto_filter.ref = worksheet_filter_ref
        workbook.save(path)

    def _run_verify(self, workbook: Path) -> tuple[subprocess.CompletedProcess[str], dict]:
        result = subprocess.run(
            [sys.executable, str(SCRIPTS / "verify_xlsx.py"), str(workbook)],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
        return result, json.loads(result.stdout)

    def test_identical_worksheet_filter_and_table_range_is_rejected(self) -> None:
        workbook = self.root / "identical-filter-overlap.xlsx"
        self._build_filter_workbook(
            workbook, worksheet_filter_ref="A1:D5", table_ref="A1:D5"
        )

        result, report = self._run_verify(workbook)

        self.assertEqual(result.returncode, 2, result.stderr)
        self.assertFalse(report["ok"])
        self.assertEqual(report["workbook"]["filter_overlap_count"], 1)
        overlap = report["workbook"]["filter_overlaps"][0]
        self.assertEqual(overlap["intersection"], "A1:D5")
        self.assertIn("ws.auto_filter.ref", overlap["message"])

    def test_partial_worksheet_filter_and_table_overlap_is_rejected(self) -> None:
        workbook = self.root / "partial-filter-overlap.xlsx"
        self._build_filter_workbook(
            workbook, worksheet_filter_ref="A1:D5", table_ref="C4:F8"
        )

        result, report = self._run_verify(workbook)

        self.assertEqual(result.returncode, 2, result.stderr)
        self.assertFalse(report["ok"])
        overlap = report["workbook"]["filter_overlaps"][0]
        self.assertEqual(overlap["worksheet_auto_filter_ref"], "A1:D5")
        self.assertEqual(overlap["table_ref"], "C4:F8")
        self.assertEqual(overlap["intersection"], "C4:D5")

    def test_disjoint_worksheet_filter_and_table_range_is_valid(self) -> None:
        workbook = self.root / "disjoint-filters.xlsx"
        self._build_filter_workbook(
            workbook, worksheet_filter_ref="A1:B4", table_ref="D1:F4"
        )

        result, report = self._run_verify(workbook)

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertTrue(report["ok"])
        self.assertEqual(report["workbook"]["filter_overlap_count"], 0)
        topology = report["workbook"]["sheets"]["Filters"]["filter_topology"]
        self.assertEqual(topology["worksheet_auto_filters"], ["A1:B4"])
        self.assertEqual(topology["tables"][0]["ref"], "D1:F4")

    def test_table_only_filter_is_valid(self) -> None:
        workbook = self.root / "table-only-filter.xlsx"
        self._build_filter_workbook(
            workbook, worksheet_filter_ref=None, table_ref="B2:E6"
        )

        result, report = self._run_verify(workbook)

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertTrue(report["ok"])
        topology = report["workbook"]["sheets"]["Filters"]["filter_topology"]
        self.assertEqual(topology["worksheet_auto_filters"], [])
        table = topology["tables"][0]
        self.assertTrue(table["relationship_id"].startswith("rId"))
        self.assertTrue(table["relationship_type"].endswith("/table"))
        self.assertEqual(table["part"], "xl/tables/table1.xml")
        self.assertEqual(table["ref"], "B2:E6")
        self.assertEqual(table["auto_filter_ref"], "B2:E6")

    def test_worksheet_filter_only_is_valid(self) -> None:
        workbook = self.root / "worksheet-filter-only.xlsx"
        self._build_filter_workbook(
            workbook, worksheet_filter_ref="A1:D5", table_ref=None
        )

        result, report = self._run_verify(workbook)

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertTrue(report["ok"])
        topology = report["workbook"]["sheets"]["Filters"]["filter_topology"]
        self.assertEqual(topology["worksheet_auto_filters"], ["A1:D5"])
        self.assertEqual(topology["tables"], [])

    def test_targeted_patch_and_policy(self) -> None:
        output = self.root / "patched.xlsx"
        spec = {
            "sheets": {
                "Template": {
                    "cells": {"B2": {"kind": "string", "value": "New value"}},
                    "row_heights": {"2": 30},
                    "data_validation": {
                        "require_count": 1,
                        "index": 0,
                        "sqref": "F2:F4",
                    },
                    "page_setup": {"fitToHeight": "1", "fitToPage": True},
                    "row_breaks": [2],
                    "print_area": "$A$1:$J$6",
                    "print_titles": "$1:$2",
                }
            }
        }
        patch_workbook(self.source, output, spec, allow_new_cells=False)

        baseline = inspect_workbook(self.source)
        current = inspect_workbook(output)
        protected = [
            name
            for name in baseline["package_entries"]
            if name.startswith("xl/drawings/")
            or name.startswith("xl/comments")
            or name.startswith("xl/media/")
            or name == "xl/styles.xml"
        ]
        policy = {
            "allowed_cells": ["Template!B2"],
            "allowed_row_heights": ["Template!2"],
            "allowed_sheet_features": [
                "Template!data_validations",
                "Template!page_setup",
                "Template!sheet_properties",
                "Template!row_breaks",
                "Workbook!defined_names",
            ],
            "required_unchanged_entries": protected,
            "allow_formula_cache_changes": False,
            "expected": {
                "formula_count": 1,
                "formula_cache_count": 0,
                "formula_error_count": 0,
            },
        }
        comparison = compare_workbooks(baseline, current, policy)
        self.assertTrue(comparison["ok"], json.dumps(comparison, indent=2))
        self.assertEqual(current["sheets"]["Template"]["cells"]["B2"]["value"], "New value")
        self.assertTrue(current["defined_names"])
        self.assertIn("$A$1:$J$6", json.dumps(current["defined_names"]))
        for name in protected:
            self.assertEqual(
                baseline["package_entries"][name], current["package_entries"][name]
            )
        with self.assertRaises(FileExistsError):
            patch_workbook(self.source, output, spec, allow_new_cells=False)

    def test_formula_cache_merge(self) -> None:
        recalc = self.root / "recalculated.xlsx"
        final = self.root / "final.xlsx"
        source_inspection = inspect_workbook(self.source)
        self.assertEqual(source_inspection["formula_count"], 1)
        self.assertEqual(source_inspection["formula_cache_count"], 0)
        entries, _, _ = load_package(self.source)
        part = resolve_sheet_parts(entries)["Template"]
        root = parse_xml(entries[part])
        formula_cell = worksheet_cells(root)["E2"]
        value = formula_cell.find(qn(MAIN_NS, "v"))
        if value is None:
            value = formula_cell.makeelement(qn(MAIN_NS, "v"), {})
            formula_cell.append(value)
        value.text = "3"
        write_package(self.source, recalc, {part: serialize_xml(root)})

        report = merge_caches(self.source, recalc, final)
        self.assertEqual(report["formula_count"], 1)
        self.assertEqual(report["formula_cache_count"], 1)
        inspected = inspect_workbook(final)
        formula = inspected["sheets"]["Template"]["cells"]["E2"]
        self.assertEqual(formula["cached"], "3")
        self.assertEqual(inspected["formula_error_count"], 0)

    def test_empty_numeric_formula_cache_is_rejected(self) -> None:
        final = self.root / "must-not-exist-empty-cache.xlsx"
        entries, _, _ = load_package(self.source)
        part = resolve_sheet_parts(entries)["Template"]
        root = parse_xml(entries[part])
        formula_cell = worksheet_cells(root)["E2"]
        self.assertFalse(has_formula_cache(formula_cell))

        with self.assertRaises(ValueError):
            merge_caches(self.source, self.source, final)
        self.assertFalse(final.exists())

    def test_empty_string_formula_cache_is_valid(self) -> None:
        recalc = self.root / "recalculated-empty-string.xlsx"
        final = self.root / "final-empty-string.xlsx"
        entries, _, _ = load_package(self.source)
        part = resolve_sheet_parts(entries)["Template"]
        root = parse_xml(entries[part])
        formula_cell = worksheet_cells(root)["E2"]
        formula_cell.attrib["t"] = "str"
        value = formula_cell.find(qn(MAIN_NS, "v"))
        if value is None:
            value = formula_cell.makeelement(qn(MAIN_NS, "v"), {})
            formula_cell.append(value)
        value.text = None
        self.assertTrue(has_formula_cache(formula_cell))
        write_package(self.source, recalc, {part: serialize_xml(root)})

        report = merge_caches(self.source, recalc, final)
        self.assertEqual(report["formula_cache_count"], 1)
        inspected = inspect_workbook(final)
        self.assertEqual(inspected["formula_cache_count"], 1)
        self.assertEqual(inspected["sheets"]["Template"]["cells"]["E2"]["cached"], "")

    def test_formula_error_cache_is_rejected(self) -> None:
        recalc = self.root / "recalculated-error.xlsx"
        final = self.root / "must-not-exist.xlsx"
        entries, _, _ = load_package(self.source)
        part = resolve_sheet_parts(entries)["Template"]
        root = parse_xml(entries[part])
        formula_cell = worksheet_cells(root)["E2"]
        value = formula_cell.find(qn(MAIN_NS, "v"))
        if value is None:
            value = formula_cell.makeelement(qn(MAIN_NS, "v"), {})
            formula_cell.append(value)
        formula_cell.attrib["t"] = "e"
        value.text = "#REF!"
        write_package(self.source, recalc, {part: serialize_xml(root)})

        with self.assertRaises(ValueError):
            merge_caches(self.source, recalc, final)
        self.assertFalse(final.exists())

    def test_policy_rejects_unapproved_cell(self) -> None:
        output = self.root / "unexpected.xlsx"
        spec = {
            "sheets": {
                "Template": {
                    "cells": {
                        "B2": {"kind": "string", "value": "Allowed"},
                        "C2": {"kind": "number", "value": 99},
                    }
                }
            }
        }
        patch_workbook(self.source, output, spec, allow_new_cells=False)
        comparison = compare_workbooks(
            inspect_workbook(self.source),
            inspect_workbook(output),
            {"allowed_cells": ["Template!B2"]},
        )
        self.assertFalse(comparison["ok"])
        self.assertEqual(comparison["unexpected"]["cells"][0]["cell"], "Template!C2")

    def test_libreoffice_recalc_when_available(self) -> None:
        if os.environ.get("RUN_LIBREOFFICE_INTEGRATION") != "1":
            self.skipTest("Set RUN_LIBREOFFICE_INTEGRATION=1 to start LibreOffice")
        try:
            soffice = find_soffice(None)
        except FileNotFoundError:
            self.skipTest("LibreOffice is not installed")
        output = self.root / "lo-recalculated.xlsx"
        convert("recalc", self.source, output, soffice, timeout=120)
        inspected = inspect_workbook(output)
        self.assertEqual(inspected["formula_count"], 1)
        self.assertEqual(inspected["formula_cache_count"], 1)
        self.assertEqual(inspected["formula_error_count"], 0)

    def test_pdf_read_only_checks(self) -> None:
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.pdfgen import canvas
        except ImportError:
            self.skipTest("reportlab is not installed")
        pdf = self.root / "sample.pdf"
        writer = canvas.Canvas(str(pdf), pagesize=landscape(A4))
        for page_number in (1, 2):
            writer.drawString(40, 40, f"Required footer - page {page_number}")
            writer.showPage()
        writer.save()
        report = inspect_pdf(
            pdf,
            expected_pages=2,
            orientation="landscape",
            expect_every_page=["Required footer"],
            expect_document=["page 2"],
            min_text_chars=1,
        )
        self.assertTrue(report["ok"], json.dumps(report, indent=2))
        self.assertTrue(report["visual_inspection_required"])

    def test_find_pdftoppm_prefers_exe_over_extensionless_wrapper(self) -> None:
        executable = self.root / "pdftoppm.exe"
        wrapper = self.root / "pdftoppm.CMD"
        executable.touch()
        wrapper.touch()

        def locate(name: str) -> str | None:
            return str(executable) if name == "pdftoppm.exe" else str(wrapper)

        with patch("verify_pdf.shutil.which", side_effect=locate):
            self.assertEqual(find_pdftoppm(None), executable.resolve())

    def test_find_pdftoppm_falls_back_to_extensionless_command(self) -> None:
        wrapper = self.root / "pdftoppm"
        wrapper.touch()

        def locate(name: str) -> str | None:
            return None if name == "pdftoppm.exe" else str(wrapper)

        with patch("verify_pdf.shutil.which", side_effect=locate):
            self.assertEqual(find_pdftoppm(None), wrapper.resolve())


class PublishOutputTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory(prefix="xlsx-publish-test-")
        self.root = Path(self.temp_dir.name)
        self.candidate = self.root / "candidate.xlsx"
        self.candidate.write_bytes(b"new workbook bytes")

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    @staticmethod
    def _sha256(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def _run_cli(self, *arguments: str) -> tuple[subprocess.CompletedProcess[str], dict]:
        result = subprocess.run(
            [sys.executable, str(SCRIPTS / "publish_output.py"), *arguments],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
        self.assertTrue(result.stdout, result.stderr)
        return result, json.loads(result.stdout)

    def _assert_no_staging_files(self, destination: Path) -> None:
        pattern = f".{destination.name}.publish-*.tmp"
        self.assertEqual(list(destination.parent.glob(pattern)), [])

    def test_create_new_destination(self) -> None:
        destination = self.root / "published.xlsx"

        result, report = self._run_cli(str(self.candidate), str(destination))

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(
            set(report),
            {
                "ok",
                "action",
                "candidate_path",
                "destination_path",
                "previous_sha256",
                "published_sha256",
            },
        )
        self.assertTrue(report["ok"])
        self.assertEqual(report["action"], "created")
        self.assertIsNone(report["previous_sha256"])
        self.assertEqual(report["published_sha256"], self._sha256(self.candidate))
        self.assertEqual(destination.read_bytes(), self.candidate.read_bytes())
        self.assertTrue(self.candidate.exists())
        self._assert_no_staging_files(destination)

    def test_replace_existing_destination_with_matching_hash(self) -> None:
        destination = self.root / "draft.xlsx"
        destination.write_bytes(b"previous draft")
        previous_sha256 = self._sha256(destination)

        result, report = self._run_cli(
            str(self.candidate),
            str(destination),
            "--replace-existing-if-sha256",
            previous_sha256,
        )

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(report["action"], "replaced")
        self.assertEqual(report["previous_sha256"], previous_sha256)
        self.assertEqual(report["published_sha256"], self._sha256(destination))
        self.assertEqual(destination.read_bytes(), self.candidate.read_bytes())
        self.assertTrue(self.candidate.exists())
        self._assert_no_staging_files(destination)

    def test_existing_destination_without_hash_is_rejected_with_failure_json(self) -> None:
        destination = self.root / "protected.xlsx"
        original = b"protected content"
        destination.write_bytes(original)

        result, report = self._run_cli(str(self.candidate), str(destination))

        self.assertEqual(result.returncode, publisher.EXIT_GUARD_REJECTED)
        self.assertEqual(
            set(report),
            {"ok", "error", "message", "candidate_path", "destination_path"},
        )
        self.assertFalse(report["ok"])
        self.assertEqual(report["error"], "destination_ownership_unconfirmed")
        self.assertEqual(destination.read_bytes(), original)

    def test_wrong_existing_hash_is_rejected(self) -> None:
        destination = self.root / "draft.xlsx"
        original = b"previous draft"
        destination.write_bytes(original)

        result, report = self._run_cli(
            str(self.candidate),
            str(destination),
            "--replace-existing-if-sha256",
            "0" * 64,
        )

        self.assertEqual(result.returncode, publisher.EXIT_GUARD_REJECTED)
        self.assertEqual(report["error"], "destination_sha256_mismatch")
        self.assertEqual(destination.read_bytes(), original)

    def test_malformed_hash_is_argument_error(self) -> None:
        destination = self.root / "draft.xlsx"
        destination.write_bytes(b"previous draft")

        result, report = self._run_cli(
            str(self.candidate),
            str(destination),
            "--replace-existing-if-sha256",
            "not-a-sha256",
        )

        self.assertEqual(result.returncode, publisher.EXIT_ARGUMENT_ERROR)
        self.assertEqual(report["error"], "invalid_sha256")

    def test_hash_is_forbidden_when_destination_is_missing(self) -> None:
        destination = self.root / "missing.xlsx"

        result, report = self._run_cli(
            str(self.candidate),
            str(destination),
            "--replace-existing-if-sha256",
            "0" * 64,
        )

        self.assertEqual(result.returncode, publisher.EXIT_ARGUMENT_ERROR)
        self.assertEqual(report["error"], "replacement_hash_without_destination")
        self.assertFalse(destination.exists())

    def test_destination_change_between_hash_checks_is_rejected(self) -> None:
        destination = self.root / "draft.xlsx"
        destination.write_bytes(b"previous draft")
        expected_sha256 = self._sha256(destination)
        real_sha256_file = publisher._sha256_file
        destination_hash_calls = 0

        def hash_then_change(path: Path) -> str:
            nonlocal destination_hash_calls
            digest = real_sha256_file(path)
            if Path(path) == destination:
                destination_hash_calls += 1
                if destination_hash_calls == 1:
                    destination.write_bytes(b"external edit")
            return digest

        with patch("publish_output._sha256_file", side_effect=hash_then_change):
            with self.assertRaises(publisher.PublishError) as caught:
                publisher.publish_output(
                    self.candidate,
                    destination,
                    replace_existing_if_sha256=expected_sha256,
                )

        self.assertEqual(caught.exception.exit_code, publisher.EXIT_GUARD_REJECTED)
        self.assertEqual(caught.exception.error, "destination_changed_before_publish")
        self.assertEqual(destination.read_bytes(), b"external edit")
        self._assert_no_staging_files(destination)

    def test_same_path_is_rejected(self) -> None:
        result, report = self._run_cli(str(self.candidate), str(self.candidate))

        self.assertEqual(result.returncode, publisher.EXIT_GUARD_REJECTED)
        self.assertEqual(report["error"], "same_path_rejected")

    def test_directory_candidate_and_destination_are_rejected(self) -> None:
        candidate_directory = self.root / "candidate-directory"
        candidate_directory.mkdir()
        destination = self.root / "published.xlsx"

        candidate_result, candidate_report = self._run_cli(
            str(candidate_directory), str(destination)
        )

        self.assertEqual(candidate_result.returncode, publisher.EXIT_GUARD_REJECTED)
        self.assertEqual(candidate_report["error"], "candidate_not_regular_file")

        destination_directory = self.root / "destination-directory"
        destination_directory.mkdir()
        destination_result, destination_report = self._run_cli(
            str(self.candidate), str(destination_directory)
        )
        self.assertEqual(destination_result.returncode, publisher.EXIT_GUARD_REJECTED)
        self.assertEqual(destination_report["error"], "destination_not_regular_file")

    def test_missing_destination_parent_is_rejected(self) -> None:
        destination = self.root / "missing-parent" / "published.xlsx"

        result, report = self._run_cli(str(self.candidate), str(destination))

        self.assertEqual(result.returncode, publisher.EXIT_GUARD_REJECTED)
        self.assertEqual(report["error"], "destination_parent_missing")
        self.assertFalse(destination.exists())

    def test_chinese_and_space_paths_are_supported(self) -> None:
        destination_directory = self.root / "中文 目录"
        destination_directory.mkdir()
        candidate = self.root / "候选 文件.xlsx"
        candidate.write_bytes(b"unicode path content")
        destination = destination_directory / "正式 文件.xlsx"

        result, report = self._run_cli(str(candidate), str(destination))

        self.assertEqual(result.returncode, 0, result.stderr)
        self.assertEqual(report["action"], "created")
        self.assertEqual(destination.read_bytes(), candidate.read_bytes())

    def test_symbolic_link_candidate_is_rejected(self) -> None:
        target = self.root / "real-candidate.xlsx"
        target.write_bytes(b"real content")
        link = self.root / "candidate-link.xlsx"
        try:
            os.symlink(target, link)
        except OSError as error:
            self.skipTest(f"symbolic links unavailable: {error}")

        result, report = self._run_cli(str(link), str(self.root / "published.xlsx"))

        self.assertEqual(result.returncode, publisher.EXIT_GUARD_REJECTED)
        self.assertEqual(report["error"], "reparse_point_rejected")

    @unittest.skipUnless(os.name == "nt", "junction test is Windows-specific")
    def test_destination_parent_junction_is_rejected(self) -> None:
        real_directory = self.root / "real-output-directory"
        real_directory.mkdir()
        junction = self.root / "output-junction"
        created = subprocess.run(
            ["cmd", "/c", "mklink", "/J", str(junction), str(real_directory)],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
        if created.returncode != 0:
            self.skipTest(f"junction creation unavailable: {created.stderr}")
        try:
            result, report = self._run_cli(
                str(self.candidate), str(junction / "published.xlsx")
            )
            self.assertEqual(result.returncode, publisher.EXIT_GUARD_REJECTED)
            self.assertEqual(report["error"], "reparse_point_rejected")
        finally:
            junction.rmdir()

    def test_copy_failure_preserves_original_and_candidate(self) -> None:
        destination = self.root / "draft.xlsx"
        original = b"original draft"
        destination.write_bytes(original)
        expected_sha256 = self._sha256(destination)

        with patch(
            "publish_output._copy_file_contents",
            side_effect=OSError(errno.EIO, "simulated copy failure"),
        ):
            with self.assertRaises(publisher.PublishError) as caught:
                publisher.publish_output(
                    self.candidate,
                    destination,
                    replace_existing_if_sha256=expected_sha256,
                )

        self.assertEqual(caught.exception.exit_code, publisher.EXIT_PUBLISH_FAILED)
        self.assertEqual(destination.read_bytes(), original)
        self.assertTrue(self.candidate.exists())
        self._assert_no_staging_files(destination)

    def test_fsync_failure_preserves_original_and_candidate(self) -> None:
        destination = self.root / "draft.xlsx"
        original = b"original draft"
        destination.write_bytes(original)
        expected_sha256 = self._sha256(destination)

        with patch(
            "publish_output.os.fsync",
            side_effect=OSError(errno.EIO, "simulated fsync failure"),
        ):
            with self.assertRaises(publisher.PublishError) as caught:
                publisher.publish_output(
                    self.candidate,
                    destination,
                    replace_existing_if_sha256=expected_sha256,
                )

        self.assertEqual(caught.exception.exit_code, publisher.EXIT_PUBLISH_FAILED)
        self.assertEqual(destination.read_bytes(), original)
        self.assertTrue(self.candidate.exists())
        self._assert_no_staging_files(destination)

    def test_locked_destination_failure_preserves_original_and_candidate(self) -> None:
        destination = self.root / "draft.xlsx"
        original = b"original draft"
        destination.write_bytes(original)
        expected_sha256 = self._sha256(destination)

        with patch(
            "publish_output._atomic_publish",
            side_effect=PermissionError(errno.EACCES, "simulated locked destination"),
        ):
            with self.assertRaises(publisher.PublishError) as caught:
                publisher.publish_output(
                    self.candidate,
                    destination,
                    replace_existing_if_sha256=expected_sha256,
                )

        self.assertEqual(caught.exception.exit_code, publisher.EXIT_PUBLISH_FAILED)
        self.assertEqual(caught.exception.error, "atomic_publish_failed")
        self.assertEqual(destination.read_bytes(), original)
        self.assertTrue(self.candidate.exists())
        self._assert_no_staging_files(destination)

    def test_staging_permission_failure_preserves_original_and_candidate(self) -> None:
        destination = self.root / "draft.xlsx"
        original = b"original draft"
        destination.write_bytes(original)
        expected_sha256 = self._sha256(destination)

        with patch(
            "publish_output.tempfile.mkstemp",
            side_effect=PermissionError(errno.EACCES, "simulated staging permission failure"),
        ):
            with self.assertRaises(publisher.PublishError) as caught:
                publisher.publish_output(
                    self.candidate,
                    destination,
                    replace_existing_if_sha256=expected_sha256,
                )

        self.assertEqual(caught.exception.exit_code, publisher.EXIT_PUBLISH_FAILED)
        self.assertEqual(caught.exception.error, "filesystem_error")
        self.assertEqual(destination.read_bytes(), original)
        self.assertTrue(self.candidate.exists())
        self._assert_no_staging_files(destination)

    def test_atomic_move_failure_preserves_original_and_candidate(self) -> None:
        destination = self.root / "draft.xlsx"
        original = b"original draft"
        destination.write_bytes(original)
        expected_sha256 = self._sha256(destination)

        with patch(
            "publish_output._atomic_publish",
            side_effect=OSError(errno.EIO, "simulated atomic move failure"),
        ):
            with self.assertRaises(publisher.PublishError) as caught:
                publisher.publish_output(
                    self.candidate,
                    destination,
                    replace_existing_if_sha256=expected_sha256,
                )

        self.assertEqual(caught.exception.exit_code, publisher.EXIT_PUBLISH_FAILED)
        self.assertEqual(destination.read_bytes(), original)
        self.assertTrue(self.candidate.exists())
        self._assert_no_staging_files(destination)

    def test_two_processes_create_same_destination_only_one_succeeds(self) -> None:
        first_candidate = self.root / "first candidate.xlsx"
        second_candidate = self.root / "second candidate.xlsx"
        first_candidate.write_bytes(b"first")
        second_candidate.write_bytes(b"second")
        destination = self.root / "raced.xlsx"
        command_prefix = [sys.executable, str(SCRIPTS / "publish_output.py")]
        processes = [
            subprocess.Popen(
                [*command_prefix, str(candidate), str(destination)],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                encoding="utf-8",
            )
            for candidate in (first_candidate, second_candidate)
        ]

        completed = []
        for process in processes:
            stdout, stderr = process.communicate(timeout=15)
            completed.append((process.returncode, json.loads(stdout), stderr))

        self.assertEqual(
            sorted(item[0] for item in completed),
            [0, publisher.EXIT_GUARD_REJECTED],
            completed,
        )
        self.assertEqual(sum(item[1]["ok"] for item in completed), 1)
        self.assertIn(destination.read_bytes(), {b"first", b"second"})
        self.assertTrue(first_candidate.exists())
        self.assertTrue(second_candidate.exists())
        self._assert_no_staging_files(destination)

    def test_main_emits_json_for_publish_failure_exit_code(self) -> None:
        destination = self.root / "published.xlsx"
        failure = publisher.PublishError(
            publisher.EXIT_PUBLISH_FAILED,
            "simulated_publish_failure",
            "simulated publish failure",
        )
        output = io.StringIO()

        with patch("publish_output.publish_output", side_effect=failure):
            with patch("sys.stdout", output):
                exit_code = publisher.main([str(self.candidate), str(destination)])

        report = json.loads(output.getvalue())
        self.assertEqual(exit_code, publisher.EXIT_PUBLISH_FAILED)
        self.assertFalse(report["ok"])
        self.assertEqual(report["error"], "simulated_publish_failure")

    def test_main_emits_json_for_unclassified_error_exit_code(self) -> None:
        destination = self.root / "published.xlsx"
        output = io.StringIO()

        with patch(
            "publish_output.publish_output",
            side_effect=RuntimeError("simulated internal failure"),
        ):
            with patch("sys.stdout", output):
                exit_code = publisher.main([str(self.candidate), str(destination)])

        report = json.loads(output.getvalue())
        self.assertEqual(exit_code, publisher.EXIT_INTERNAL_ERROR)
        self.assertFalse(report["ok"])
        self.assertEqual(report["error"], "internal_error")


class SkillStructureTests(unittest.TestCase):
    def test_required_files_exist(self) -> None:
        required = [
            "SKILL.md",
            "references/general-workflow.md",
            "references/formatting-and-formulas.md",
            "references/high-fidelity-workflow.md",
            "references/patch-spec.md",
            "references/output-lifecycle.md",
            "scripts/patch_ooxml.py",
            "scripts/libreoffice_headless.py",
            "scripts/merge_formula_caches.py",
            "scripts/verify_xlsx.py",
            "scripts/verify_pdf.py",
            "scripts/publish_output.py",
            "evals/evals.json",
        ]
        for relative in required:
            self.assertTrue((SKILL_ROOT / relative).is_file(), relative)

    def test_skill_is_complete_xlsx_entry(self) -> None:
        skill_text = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")
        workflow_text = (SKILL_ROOT / "references" / "general-workflow.md").read_text(
            encoding="utf-8"
        )
        self.assertIn("name: xlsx", skill_text)
        self.assertIn("这是完整的表格技能", skill_text)
        self.assertIn("ws.auto_filter.ref", skill_text)
        self.assertIn("ws.auto_filter.ref", workflow_text)
        self.assertNotIn("xlsx-" + "preserve-ooxml", skill_text)
        self.assertNotIn("不使用本 " + "skill", skill_text)

    def test_eval_json_is_valid(self) -> None:
        data = json.loads((SKILL_ROOT / "evals" / "evals.json").read_text(encoding="utf-8"))
        self.assertEqual(data["skill_name"], "xlsx")
        self.assertGreaterEqual(len(data["evals"]), 9)
        lifecycle_eval = next(item for item in data["evals"] if item["id"] == 9)
        lifecycle_text = json.dumps(lifecycle_eval, ensure_ascii=False)
        self.assertIn("交付前", lifecycle_text)
        self.assertIn("SHA-256", lifecycle_text)
        self.assertIn("递增版本", lifecycle_text)

        triggers = json.loads(
            (SKILL_ROOT / "evals" / "trigger-evals.json").read_text(encoding="utf-8")
        )
        self.assertEqual(len(triggers), 20)
        self.assertEqual(sum(item["should_trigger"] for item in triggers), 10)


if __name__ == "__main__":
    unittest.main()
