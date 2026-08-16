from __future__ import annotations

import hashlib
import json
import math
import os
import re
import shutil
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import urlparse

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


SCHEMA_VERSION = "1.0"
SUPPORTED_OPENPYXL_VERSION = "3.1.5"

FIELD_DESIGN_SHEET = "字段设计"
PRODUCT_CANDIDATES_SHEET = "产品候选"
PARAMETER_EVIDENCE_SHEET = "参数证据"
PROCUREMENT_EVIDENCE_SHEET = "采购证据"
PROCUREMENT_FORMAL_SHEET = "采购渠道"

FIELD_DESIGN_HEADERS = (
    "工作表顺序",
    "工作表名称",
    "字段顺序",
    "字段名称",
    "单位",
    "字段角色",
    "值类型",
    "确认状态",
    "样本依据",
)
PRODUCT_CANDIDATE_HEADERS = (
    "候选编号",
    "正式产品编号",
    "目标参数表",
    "国家",
    "产品名称",
    "产品型号",
    "品牌",
    "生产厂家",
    "纳入状态",
    "复核备注",
)
PARAMETER_EVIDENCE_HEADERS = (
    "证据编号",
    "候选编号",
    "字段名称",
    "原始值",
    "正式值",
    "来源类型",
    "来源标题",
    "来源URL",
    "支持摘录",
    "页码/章节",
    "访问日期",
    "证据状态",
    "正式采用",
    "冲突说明",
)
PROCUREMENT_EVIDENCE_HEADERS = (
    "渠道记录编号",
    "候选编号",
    "渠道公司",
    "渠道角色",
    "可销售地区",
    "公开价格",
    "最小起订量",
    "电话",
    "邮箱",
    "询价或购买链接",
    "国内代理/销售链接",
    "采购或供货说明",
    "资料来源",
    "来源标题",
    "支持摘录",
    "访问日期",
    "证据状态",
    "正式采用",
    "冲突说明",
)
PROCESS_HEADERS = {
    FIELD_DESIGN_SHEET: FIELD_DESIGN_HEADERS,
    PRODUCT_CANDIDATES_SHEET: PRODUCT_CANDIDATE_HEADERS,
    PARAMETER_EVIDENCE_SHEET: PARAMETER_EVIDENCE_HEADERS,
    PROCUREMENT_EVIDENCE_SHEET: PROCUREMENT_EVIDENCE_HEADERS,
}
PROCESS_TABLE_NAMES = {
    FIELD_DESIGN_SHEET: "PRW_FieldDesign",
    PRODUCT_CANDIDATES_SHEET: "PRW_ProductCandidates",
    PARAMETER_EVIDENCE_SHEET: "PRW_ParameterEvidence",
    PROCUREMENT_EVIDENCE_SHEET: "PRW_ProcurementEvidence",
}

IDENTITY_FIELDS = (
    "产品编号",
    "国家",
    "产品名称",
    "产品型号",
    "品牌",
    "生产厂家",
)
CANDIDATE_IDENTITY_FIELDS = IDENTITY_FIELDS[1:]
PARAMETER_TRAILING_FIELDS = ("产品页", "规格书/手册", "备注")
PROCUREMENT_BUSINESS_FIELDS = (
    "渠道公司",
    "渠道角色",
    "可销售地区",
    "公开价格",
    "最小起订量",
    "电话",
    "邮箱",
    "询价或购买链接",
    "国内代理/销售链接",
    "采购或供货说明",
    "资料来源",
)
PROCUREMENT_FORMAL_HEADERS = (
    "产品编号",
    "产品名称",
    "产品型号",
    *PROCUREMENT_BUSINESS_FIELDS,
)
REQUIRED_HYPERLINK_FIELDS = {
    "产品页",
    "规格书/手册",
    "询价或购买链接",
    "国内代理/销售链接",
}

FIELD_ROLES = ("身份", "技术参数", "参数来源", "备注")
VALUE_TYPES = ("文本", "整数", "小数", "日期", "布尔值", "URL")
FIELD_CONFIRMATION_STATES = ("候选", "已确认")
INCLUSION_STATES = ("待核验", "待审核", "已确认纳入", "已排除")
EVIDENCE_STATES = ("待复核", "已确认", "冲突")
ADOPTION_STATES = ("是", "否")
SOURCE_TYPES = (
    "产品页",
    "规格书/手册",
    "其他官方资料",
    "授权渠道页",
    "其他公开资料",
)

FORBIDDEN_FORMAL_HEADERS = {
    "证据等级",
    "证据原文",
    "证据摘录",
    "支持摘录",
    "判定依据",
    "采购状态",
    "采购流程状态",
    "访问日期",
    "证据状态",
    "正式采用",
    "冲突说明",
}
FORBIDDEN_PLACEHOLDERS = {
    "未说明",
    "未公开",
    "待核实",
    "未知",
    "N/A",
    "NA",
}

CANDIDATE_ID_PATTERN = re.compile(r"^CAND-\d{4,}$")
SHEET_FORBIDDEN_PATTERN = re.compile(r"[\\/*?:\[\]]")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
UNIT_FILL = PatternFill("solid", fgColor="D9EAF7")
UNIT_FONT = Font(name="Aptos", size=9, italic=True, color="244062")
BODY_FONT = Font(name="Aptos", size=10, color="1F1F1F")
LINK_FONT = Font(name="Aptos", size=10, color="0563C1", underline="single")
THIN_GRAY = Side(style="thin", color="B7C9D6")
CELL_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def ensure_supported_openpyxl() -> None:
    if openpyxl.__version__ != SUPPORTED_OPENPYXL_VERSION:
        raise RuntimeError(
            f"openpyxl {SUPPORTED_OPENPYXL_VERSION} is required; found {openpyxl.__version__}"
        )


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def canonical_scalar(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return str(value)
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    return clean_text(value)


def parse_positive_int(value: Any, *, require_numeric: bool = True) -> int | None:
    if is_blank(value) or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        number = int(value)
        return number if number > 0 else None
    if not require_numeric and isinstance(value, str) and value.strip().isdigit():
        number = int(value.strip())
        return number if number > 0 else None
    return None


def is_http_url(value: Any) -> bool:
    if is_blank(value):
        return False
    parsed = urlparse(clean_text(value))
    return parsed.scheme.lower() in {"http", "https"} and bool(parsed.netloc)


def is_valid_sheet_name(value: Any) -> bool:
    name = clean_text(value)
    return (
        bool(name)
        and len(name) <= 31
        and not SHEET_FORBIDDEN_PATTERN.search(name)
        and not (name.startswith("'") or name.endswith("'"))
    )


def coerce_formal_value(value: Any, value_type: str) -> Any:
    if is_blank(value):
        return None
    text = clean_text(value)
    if value_type in {"文本", "URL"}:
        return text
    if value_type == "整数":
        if not re.fullmatch(r"[+-]?\d+", text):
            raise ValueError(f"cannot parse integer value: {text}")
        return int(text)
    if value_type == "小数":
        number = float(text)
        if not math.isfinite(number):
            raise ValueError(f"non-finite decimal value: {text}")
        return number
    if value_type == "日期":
        return date.fromisoformat(text)
    if value_type == "布尔值":
        mapping = {"是": True, "否": False, "true": True, "false": False}
        if text not in mapping:
            raise ValueError(f"cannot parse boolean value: {text}")
        return mapping[text]
    raise ValueError(f"unsupported value type: {value_type}")


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def publish_file_no_overwrite(source: str | Path, destination: str | Path) -> str:
    source_path = Path(source)
    destination_path = Path(destination)
    if destination_path.exists():
        raise FileExistsError(f"output already exists: {destination_path}")
    if not destination_path.parent.is_dir():
        raise FileNotFoundError(f"output directory does not exist: {destination_path.parent}")

    created = False
    try:
        with source_path.open("rb") as source_handle, destination_path.open("xb") as output_handle:
            created = True
            shutil.copyfileobj(source_handle, output_handle, length=1024 * 1024)
            output_handle.flush()
            os.fsync(output_handle.fileno())
    except Exception:
        if created:
            destination_path.unlink(missing_ok=True)
        raise
    return sha256_file(destination_path)


def save_workbook_no_overwrite(workbook: Workbook, destination: str | Path) -> str:
    destination_path = Path(destination)
    if destination_path.suffix.lower() != ".xlsx":
        raise ValueError("output must use the .xlsx extension")
    if destination_path.exists():
        raise FileExistsError(f"output already exists: {destination_path}")
    if not destination_path.parent.is_dir():
        raise FileNotFoundError(f"output directory does not exist: {destination_path.parent}")

    file_descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{destination_path.stem}-", suffix=".xlsx", dir=destination_path.parent
    )
    os.close(file_descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        return publish_file_no_overwrite(temporary_path, destination_path)
    finally:
        temporary_path.unlink(missing_ok=True)


def write_json_no_overwrite(payload: dict[str, Any], destination: str | Path) -> str:
    destination_path = Path(destination)
    if destination_path.exists():
        raise FileExistsError(f"output already exists: {destination_path}")
    if not destination_path.parent.is_dir():
        raise FileNotFoundError(f"output directory does not exist: {destination_path.parent}")
    encoded = (json.dumps(payload, ensure_ascii=False, indent=2) + "\n").encode("utf-8")
    created = False
    try:
        with destination_path.open("xb") as handle:
            created = True
            handle.write(encoded)
            handle.flush()
            os.fsync(handle.fileno())
    except Exception:
        if created:
            destination_path.unlink(missing_ok=True)
        raise
    return hashlib.sha256(encoded).hexdigest()


def last_nonblank_column(values: Iterable[Any]) -> int:
    result = 0
    for index, value in enumerate(values, start=1):
        if not is_blank(value):
            result = index
    return result


def last_nonblank_row(worksheet: Any, start_row: int = 1) -> int:
    result = start_row - 1
    for row in worksheet.iter_rows(min_row=start_row):
        if any(not is_blank(cell.value) for cell in row):
            result = row[0].row
    return result


def ranges_intersect(first: str, second: str) -> bool:
    first_min_col, first_min_row, first_max_col, first_max_row = range_boundaries(first)
    second_min_col, second_min_row, second_max_col, second_max_row = range_boundaries(second)
    return not (
        first_max_col < second_min_col
        or second_max_col < first_min_col
        or first_max_row < second_min_row
        or second_max_row < first_min_row
    )


def add_external_hyperlink(cell: Any, target: str) -> None:
    cell.hyperlink = target
    cell.font = LINK_FONT


def apply_data_validation(worksheet: Any, column: int, values: Iterable[str]) -> None:
    formula = '"' + ",".join(values) + '"'
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "请选择契约允许的值"
    validation.errorTitle = "值不符合工作簿契约"
    worksheet.add_data_validation(validation)
    letter = get_column_letter(column)
    validation.add(f"{letter}2:{letter}10000")


def _set_process_widths(worksheet: Any, headers: tuple[str, ...]) -> None:
    long_fields = {
        "样本依据",
        "复核备注",
        "原始值",
        "正式值",
        "来源标题",
        "来源URL",
        "支持摘录",
        "冲突说明",
        "采购或供货说明",
        "资料来源",
    }
    for index, header in enumerate(headers, start=1):
        width = 34 if header in long_fields else 18
        worksheet.column_dimensions[get_column_letter(index)].width = width


def create_process_workbook() -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)

    for sheet_name, headers in PROCESS_HEADERS.items():
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.append(list(headers))
        worksheet.append([None] * len(headers))
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        worksheet.row_dimensions[1].height = 30
        _set_process_widths(worksheet, headers)

        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = CELL_BORDER
        for cell in worksheet[2]:
            cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT
            cell.border = CELL_BORDER

        last_column = get_column_letter(len(headers))
        table = Table(
            displayName=PROCESS_TABLE_NAMES[sheet_name],
            ref=f"A1:{last_column}2",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    design = workbook[FIELD_DESIGN_SHEET]
    apply_data_validation(design, FIELD_DESIGN_HEADERS.index("字段角色") + 1, FIELD_ROLES)
    apply_data_validation(design, FIELD_DESIGN_HEADERS.index("值类型") + 1, VALUE_TYPES)
    apply_data_validation(
        design,
        FIELD_DESIGN_HEADERS.index("确认状态") + 1,
        FIELD_CONFIRMATION_STATES,
    )

    candidates = workbook[PRODUCT_CANDIDATES_SHEET]
    apply_data_validation(
        candidates,
        PRODUCT_CANDIDATE_HEADERS.index("纳入状态") + 1,
        INCLUSION_STATES,
    )

    parameter_evidence = workbook[PARAMETER_EVIDENCE_SHEET]
    apply_data_validation(
        parameter_evidence,
        PARAMETER_EVIDENCE_HEADERS.index("来源类型") + 1,
        SOURCE_TYPES,
    )
    apply_data_validation(
        parameter_evidence,
        PARAMETER_EVIDENCE_HEADERS.index("证据状态") + 1,
        EVIDENCE_STATES,
    )
    apply_data_validation(
        parameter_evidence,
        PARAMETER_EVIDENCE_HEADERS.index("正式采用") + 1,
        ADOPTION_STATES,
    )

    procurement_evidence = workbook[PROCUREMENT_EVIDENCE_SHEET]
    apply_data_validation(
        procurement_evidence,
        PROCUREMENT_EVIDENCE_HEADERS.index("证据状态") + 1,
        EVIDENCE_STATES,
    )
    apply_data_validation(
        procurement_evidence,
        PROCUREMENT_EVIDENCE_HEADERS.index("正式采用") + 1,
        ADOPTION_STATES,
    )

    return workbook


def style_formal_parameter_sheet(worksheet: Any, column_count: int) -> None:
    worksheet.freeze_panes = "A3"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 34
    worksheet.row_dimensions[2].height = 24
    worksheet.print_title_rows = "1:2"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.5
    worksheet.page_margins.bottom = 0.5

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = CELL_BORDER
    for cell in worksheet[2]:
        cell.fill = UNIT_FILL
        cell.font = UNIT_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = CELL_BORDER

    for column in range(1, column_count + 1):
        header = clean_text(worksheet.cell(1, column).value)
        if header == "产品编号":
            width = 11
        elif header in REQUIRED_HYPERLINK_FIELDS or header == "备注":
            width = 32
        elif header in {"产品名称", "产品型号", "生产厂家"}:
            width = 22
        else:
            width = 16
        worksheet.column_dimensions[get_column_letter(column)].width = width


def style_formal_procurement_sheet(worksheet: Any) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 34
    worksheet.print_title_rows = "1:1"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    widths = (11, 22, 20, 20, 16, 16, 16, 14, 16, 18, 30, 30, 34, 28)
    for index, width in enumerate(widths, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = CELL_BORDER


def style_formal_body(worksheet: Any, start_row: int) -> None:
    for row in worksheet.iter_rows(min_row=start_row):
        worksheet.row_dimensions[row[0].row].height = 34
        for cell in row:
            if cell.hyperlink is None:
                cell.font = BODY_FONT
            cell.alignment = WRAP_ALIGNMENT
            cell.border = CELL_BORDER
        if row:
            row[0].alignment = CENTER_ALIGNMENT


def add_procurement_table(worksheet: Any) -> None:
    last_row = max(2, worksheet.max_row)
    last_column = get_column_letter(len(PROCUREMENT_FORMAL_HEADERS))
    table = Table(displayName="PRW_FormalProcurement", ref=f"A1:{last_column}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def add_missing_value_highlight(worksheet: Any, cell_range: str) -> None:
    fill = PatternFill("solid", fgColor="FFF2CC")
    worksheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'LEN(TRIM({cell_range.split(":")[0]}))=0'], fill=fill),
    )
