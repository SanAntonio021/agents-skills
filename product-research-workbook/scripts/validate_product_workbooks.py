from __future__ import annotations

import argparse
import json
import posixpath
import sys
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries

from workbook_contract import (
    ADOPTION_STATES,
    CANDIDATE_ID_PATTERN,
    CANDIDATE_IDENTITY_FIELDS,
    EVIDENCE_STATES,
    FIELD_CONFIRMATION_STATES,
    FIELD_DESIGN_HEADERS,
    FIELD_DESIGN_SHEET,
    FIELD_ROLES,
    FORBIDDEN_FORMAL_HEADERS,
    FORBIDDEN_PLACEHOLDERS,
    IDENTITY_FIELDS,
    INCLUSION_STATES,
    PARAMETER_EVIDENCE_HEADERS,
    PARAMETER_EVIDENCE_SHEET,
    PARAMETER_TRAILING_FIELDS,
    PROCESS_HEADERS,
    PROCUREMENT_BUSINESS_FIELDS,
    PROCUREMENT_EVIDENCE_HEADERS,
    PROCUREMENT_EVIDENCE_SHEET,
    PROCUREMENT_FORMAL_HEADERS,
    PROCUREMENT_FORMAL_SHEET,
    PRODUCT_CANDIDATE_HEADERS,
    PRODUCT_CANDIDATES_SHEET,
    REQUIRED_HYPERLINK_FIELDS,
    SCHEMA_VERSION,
    SOURCE_TYPES,
    VALUE_TYPES,
    canonical_scalar,
    clean_text,
    coerce_formal_value,
    ensure_supported_openpyxl,
    is_blank,
    is_http_url,
    is_valid_sheet_name,
    last_nonblank_column,
    last_nonblank_row,
    parse_positive_int,
    ranges_intersect,
    sha256_file,
    write_json_no_overwrite,
)


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


@dataclass
class Finding:
    code: str
    severity: str
    workbook: str | None
    sheet: str | None
    cell_or_row: str | None
    candidate_id: str | None
    product_id: int | None
    message: str
    evidence: Any = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "code": self.code,
            "severity": self.severity,
            "workbook": self.workbook,
            "sheet": self.sheet,
            "cell_or_row": self.cell_or_row,
            "candidate_id": self.candidate_id,
            "product_id": self.product_id,
            "message": self.message,
            "evidence": self.evidence,
        }


@dataclass
class ValidationReport:
    mode: str
    inputs: dict[str, dict[str, str] | None]
    findings: list[Finding] = field(default_factory=list)
    product_count: int = 0
    channel_count: int = 0

    def add(
        self,
        code: str,
        message: str,
        *,
        severity: str = "error",
        workbook: str | None = None,
        sheet: str | None = None,
        cell_or_row: str | None = None,
        candidate_id: str | None = None,
        product_id: int | None = None,
        evidence: Any = None,
    ) -> None:
        self.findings.append(
            Finding(
                code=code,
                severity=severity,
                workbook=workbook,
                sheet=sheet,
                cell_or_row=cell_or_row,
                candidate_id=candidate_id,
                product_id=product_id,
                message=message,
                evidence=evidence,
            )
        )

    @property
    def error_count(self) -> int:
        return sum(item.severity == "error" for item in self.findings)

    @property
    def warning_count(self) -> int:
        return sum(item.severity == "warning" for item in self.findings)

    @property
    def ok(self) -> bool:
        return self.error_count == 0

    def to_dict(self) -> dict[str, Any]:
        return {
            "schema_version": SCHEMA_VERSION,
            "ok": self.ok,
            "mode": self.mode,
            "inputs": self.inputs,
            "summary": {
                "product_count": self.product_count,
                "channel_count": self.channel_count,
                "error_count": self.error_count,
                "warning_count": self.warning_count,
            },
            "findings": [item.to_dict() for item in self.findings],
        }


@dataclass(frozen=True)
class RowRecord:
    row: int
    values: dict[str, Any]


@dataclass(frozen=True)
class FieldRecord:
    row: int
    sheet_order: int
    sheet_name: str
    field_order: int
    field_name: str
    unit: str
    role: str
    value_type: str
    confirmation: str
    sample_basis: str


@dataclass(frozen=True)
class CandidateRecord:
    row: int
    candidate_id: str
    product_id: int | None
    target_sheet: str
    identities: dict[str, Any]
    inclusion: str


@dataclass(frozen=True)
class ParameterEvidenceRecord:
    row: int
    evidence_id: str
    candidate_id: str
    field_name: str
    original_value: Any
    formal_value: Any
    source_type: str
    source_title: str
    source_url: str
    excerpt: str
    access_date: Any
    evidence_state: str
    adopted: str
    conflict_note: str


@dataclass(frozen=True)
class ProcurementEvidenceRecord:
    row: int
    channel_id: str
    candidate_id: str
    business_values: dict[str, Any]
    source_title: str
    excerpt: str
    access_date: Any
    evidence_state: str
    adopted: str
    conflict_note: str


@dataclass
class ProcessContext:
    field_records: list[FieldRecord]
    fields_by_sheet: dict[str, list[FieldRecord]]
    confirmed_fields_by_sheet: dict[str, list[FieldRecord]]
    candidates: list[CandidateRecord]
    candidates_by_id: dict[str, CandidateRecord]
    candidates_by_product_id: dict[int, CandidateRecord]
    parameter_evidence: list[ParameterEvidenceRecord]
    main_evidence: dict[tuple[str, str], ParameterEvidenceRecord]
    procurement_evidence: list[ProcurementEvidenceRecord]
    adopted_procurement: list[ProcurementEvidenceRecord]


@dataclass
class SheetXmlInfo:
    hyperlinks: list[dict[str, str | None]] = field(default_factory=list)
    worksheet_auto_filter: str | None = None
    table_refs: list[str] = field(default_factory=list)


def _input_descriptor(path: str | Path | None) -> dict[str, str] | None:
    if path is None:
        return None
    resolved = Path(path).resolve()
    return {"path": str(resolved), "sha256": sha256_file(resolved)}


def _column_cell(headers: tuple[str, ...], field_name: str, row: int) -> str:
    return f"{get_column_letter(headers.index(field_name) + 1)}{row}"


def _formal_product_id_error(value: Any, subject: str) -> tuple[str, str]:
    if isinstance(value, str) and clean_text(value):
        return (
            "FORMAL_PRODUCT_ID_TEXT_FORBIDDEN",
            f"{subject}必须为 Excel 数值型正整数；不得填入历史人工前缀序号、候选编号或厂家型号。",
        )
    return "FORMAL_PRODUCT_ID_INVALID", f"{subject}必须为空或为正整数数值。"


def _sheet_rows(worksheet: Any, headers: tuple[str, ...]) -> list[RowRecord]:
    rows: list[RowRecord] = []
    for row_number in range(2, worksheet.max_row + 1):
        values = {
            header: worksheet.cell(row_number, column_number).value
            for column_number, header in enumerate(headers, start=1)
        }
        if any(not is_blank(value) for value in values.values()):
            rows.append(RowRecord(row=row_number, values=values))
    return rows


def _validate_no_formulas(workbook: Any, report: ValidationReport, workbook_name: str) -> None:
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    report.add(
                        "FORMULA_NOT_ALLOWED",
                        "产品调研过程表和正式表不允许公式。",
                        workbook=workbook_name,
                        sheet=worksheet.title,
                        cell_or_row=cell.coordinate,
                        evidence={"formula": cell.value},
                    )


def _validate_process_sheet_structure(
    workbook: Any, report: ValidationReport
) -> dict[str, list[RowRecord]]:
    parsed: dict[str, list[RowRecord]] = {}
    for sheet_name, expected_headers in PROCESS_HEADERS.items():
        if sheet_name not in workbook.sheetnames:
            report.add(
                "PROCESS_SHEET_MISSING",
                f"缺少过程工作表：{sheet_name}",
                workbook="process",
                sheet=sheet_name,
            )
            parsed[sheet_name] = []
            continue

        worksheet = workbook[sheet_name]
        actual_width = last_nonblank_column(
            worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)
        )
        actual_headers = tuple(
            clean_text(worksheet.cell(1, column).value)
            for column in range(1, actual_width + 1)
        )
        if actual_headers != expected_headers:
            report.add(
                "PROCESS_HEADER_MISMATCH",
                "过程工作表表头必须与契约完全一致。",
                workbook="process",
                sheet=sheet_name,
                cell_or_row="1",
                evidence={"expected": list(expected_headers), "actual": list(actual_headers)},
            )

        if worksheet.freeze_panes != "A2":
            report.add(
                "PROCESS_FREEZE_PANE_INVALID",
                "过程工作表必须冻结首行（A2）。",
                workbook="process",
                sheet=sheet_name,
                evidence={"actual": str(worksheet.freeze_panes)},
            )

        if worksheet.auto_filter.ref:
            report.add(
                "WORKSHEET_AUTOFILTER_FORBIDDEN",
                "过程工作表使用 Excel Table 自带筛选，不得另设工作表级 autoFilter。",
                workbook="process",
                sheet=sheet_name,
                evidence={"ref": worksheet.auto_filter.ref},
            )

        tables = list(worksheet.tables.values())
        last_data_row = max(2, last_nonblank_row(worksheet, 2))
        expected_last_column = len(expected_headers)
        if len(tables) != 1:
            report.add(
                "PROCESS_TABLE_INVALID",
                "每张过程工作表必须且只能包含一个 Excel Table。",
                workbook="process",
                sheet=sheet_name,
                evidence={"table_count": len(tables)},
            )
        else:
            min_col, min_row, max_col, max_row = range_boundaries(tables[0].ref)
            if (
                min_col != 1
                or min_row != 1
                or max_col != expected_last_column
                or max_row < last_data_row
                or max_row < 2
            ):
                report.add(
                    "PROCESS_TABLE_RANGE_INVALID",
                    "过程表范围必须从 A1 开始，覆盖完整表头和全部数据行。",
                    workbook="process",
                    sheet=sheet_name,
                    evidence={"table_ref": tables[0].ref, "last_data_row": last_data_row},
                )

        parsed[sheet_name] = _sheet_rows(worksheet, expected_headers)
    return parsed


def _value_in(value: str, allowed: Iterable[str]) -> bool:
    return value in set(allowed)


def _forbidden_placeholder(value: Any) -> bool:
    return clean_text(value).upper() in {item.upper() for item in FORBIDDEN_PLACEHOLDERS}


def _parse_fields(
    rows: list[RowRecord], report: ValidationReport
) -> tuple[list[FieldRecord], dict[str, list[FieldRecord]], dict[str, list[FieldRecord]]]:
    records: list[FieldRecord] = []
    seen_fields: set[tuple[str, int]] = set()
    seen_names: set[tuple[str, str]] = set()
    sheet_orders: dict[str, int] = {}
    order_to_sheet: dict[int, str] = {}

    for item in rows:
        values = item.values
        sheet_name = clean_text(values["工作表名称"])
        field_name = clean_text(values["字段名称"])
        sheet_order = parse_positive_int(values["工作表顺序"])
        field_order = parse_positive_int(values["字段顺序"])
        unit = clean_text(values["单位"])
        role = clean_text(values["字段角色"])
        value_type = clean_text(values["值类型"])
        confirmation = clean_text(values["确认状态"])
        sample_basis = clean_text(values["样本依据"])

        if sheet_order is None:
            report.add(
                "FIELD_SHEET_ORDER_INVALID",
                "工作表顺序必须是正整数。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "工作表顺序", item.row),
                evidence={"value": values["工作表顺序"]},
            )
        if field_order is None:
            report.add(
                "FIELD_ORDER_INVALID",
                "字段顺序必须是正整数。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "字段顺序", item.row),
                evidence={"value": values["字段顺序"]},
            )
        if not is_valid_sheet_name(sheet_name) or sheet_name in {
            *PROCESS_HEADERS.keys(),
            PROCUREMENT_FORMAL_SHEET,
        }:
            report.add(
                "PARAMETER_SHEET_NAME_INVALID",
                "参数工作表名称为空、含 Excel 禁用字符、超过 31 字符或与保留名称冲突。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "工作表名称", item.row),
                evidence={"value": sheet_name},
            )
        if not field_name:
            report.add(
                "FIELD_NAME_MISSING",
                "字段名称不能为空。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "字段名称", item.row),
            )
        if role not in FIELD_ROLES:
            report.add(
                "FIELD_ROLE_INVALID",
                "字段角色不在契约枚举中。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "字段角色", item.row),
                evidence={"value": role, "allowed": list(FIELD_ROLES)},
            )
        if value_type not in VALUE_TYPES:
            report.add(
                "FIELD_VALUE_TYPE_INVALID",
                "值类型不在契约枚举中。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "值类型", item.row),
                evidence={"value": value_type, "allowed": list(VALUE_TYPES)},
            )
        if confirmation not in FIELD_CONFIRMATION_STATES:
            report.add(
                "FIELD_CONFIRMATION_INVALID",
                "确认状态必须为“候选”或“已确认”。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "确认状态", item.row),
                evidence={"value": confirmation},
            )
        if role != "技术参数" and unit:
            report.add(
                "UNIT_ON_NONTECHNICAL_FIELD",
                "身份、来源和备注字段的单位必须留空。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "单位", item.row),
                evidence={"field": field_name, "unit": unit},
            )
        if role == "技术参数" and not sample_basis:
            report.add(
                "SAMPLE_BASIS_REQUIRED",
                "技术参数字段必须记录形成该字段的真实样本依据。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "样本依据", item.row),
                evidence={"field": field_name},
            )
        if unit and unit.casefold() in field_name.casefold():
            report.add(
                "FIELD_NAME_CONTAINS_UNIT",
                "字段名不得重复单位；单位只写在单位列和正式表第二行。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=_column_cell(FIELD_DESIGN_HEADERS, "字段名称", item.row),
                evidence={"field": field_name, "unit": unit},
            )

        if sheet_order is None or field_order is None or not sheet_name or not field_name:
            continue

        order_key = (sheet_name, field_order)
        name_key = (sheet_name, field_name)
        if order_key in seen_fields:
            report.add(
                "FIELD_ORDER_DUPLICATE",
                "同一参数表内字段顺序不得重复。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=str(item.row),
                evidence={"sheet": sheet_name, "field_order": field_order},
            )
        seen_fields.add(order_key)
        if name_key in seen_names:
            report.add(
                "FIELD_NAME_DUPLICATE",
                "同一参数表内字段名称不得重复。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=str(item.row),
                evidence={"sheet": sheet_name, "field": field_name},
            )
        seen_names.add(name_key)

        if sheet_name in sheet_orders and sheet_orders[sheet_name] != sheet_order:
            report.add(
                "SHEET_ORDER_INCONSISTENT",
                "同一参数表的工作表顺序必须一致。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=str(item.row),
                evidence={"sheet": sheet_name, "orders": [sheet_orders[sheet_name], sheet_order]},
            )
        sheet_orders.setdefault(sheet_name, sheet_order)
        if sheet_order in order_to_sheet and order_to_sheet[sheet_order] != sheet_name:
            report.add(
                "SHEET_ORDER_DUPLICATE",
                "不同参数表不得使用同一工作表顺序。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=str(item.row),
                evidence={"order": sheet_order, "sheets": [order_to_sheet[sheet_order], sheet_name]},
            )
        order_to_sheet.setdefault(sheet_order, sheet_name)

        records.append(
            FieldRecord(
                row=item.row,
                sheet_order=sheet_order,
                sheet_name=sheet_name,
                field_order=field_order,
                field_name=field_name,
                unit=unit,
                role=role,
                value_type=value_type,
                confirmation=confirmation,
                sample_basis=sample_basis,
            )
        )

    fields_by_sheet: dict[str, list[FieldRecord]] = defaultdict(list)
    confirmed_by_sheet: dict[str, list[FieldRecord]] = defaultdict(list)
    for record in records:
        fields_by_sheet[record.sheet_name].append(record)
        if record.confirmation == "已确认":
            confirmed_by_sheet[record.sheet_name].append(record)
    for values in fields_by_sheet.values():
        values.sort(key=lambda record: record.field_order)
    for values in confirmed_by_sheet.values():
        values.sort(key=lambda record: record.field_order)

    for sheet_name, fields in confirmed_by_sheet.items():
        names = tuple(record.field_name for record in fields)
        expected_minimum = len(IDENTITY_FIELDS) + len(PARAMETER_TRAILING_FIELDS)
        if len(names) < expected_minimum:
            report.add(
                "CONFIRMED_FIELD_LAYOUT_INVALID",
                "已确认参数表缺少固定身份字段或固定尾部来源字段。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                evidence={"parameter_sheet": sheet_name, "fields": list(names)},
            )
            continue
        if names[: len(IDENTITY_FIELDS)] != IDENTITY_FIELDS or names[-3:] != PARAMETER_TRAILING_FIELDS:
            report.add(
                "CONFIRMED_FIELD_LAYOUT_INVALID",
                "参数表前六列和末三列必须采用固定顺序。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                evidence={
                    "parameter_sheet": sheet_name,
                    "expected_prefix": list(IDENTITY_FIELDS),
                    "expected_suffix": list(PARAMETER_TRAILING_FIELDS),
                    "actual": list(names),
                },
            )
        for index, record in enumerate(fields):
            if index < len(IDENTITY_FIELDS):
                expected_role = "身份"
            elif index >= len(fields) - len(PARAMETER_TRAILING_FIELDS):
                expected_role = "备注" if record.field_name == "备注" else "参数来源"
            else:
                expected_role = "技术参数"
            if record.role != expected_role:
                report.add(
                    "CONFIRMED_FIELD_ROLE_INVALID",
                    "字段角色与其在正式参数表中的位置不一致。",
                    workbook="process",
                    sheet=FIELD_DESIGN_SHEET,
                    cell_or_row=str(record.row),
                    evidence={"field": record.field_name, "expected": expected_role, "actual": record.role},
                )
        if fields and fields[0].value_type != "整数":
            report.add(
                "PRODUCT_ID_TYPE_INVALID",
                "产品编号字段的值类型必须为整数。",
                workbook="process",
                sheet=FIELD_DESIGN_SHEET,
                cell_or_row=str(fields[0].row),
            )
        for record in fields:
            if record.field_name in {"产品页", "规格书/手册"} and record.value_type != "URL":
                report.add(
                    "SOURCE_FIELD_TYPE_INVALID",
                    "产品页和规格书/手册字段的值类型必须为 URL。",
                    workbook="process",
                    sheet=FIELD_DESIGN_SHEET,
                    cell_or_row=str(record.row),
                    evidence={"field": record.field_name, "actual": record.value_type},
                )

    return records, dict(fields_by_sheet), dict(confirmed_by_sheet)


def _parse_candidates(
    rows: list[RowRecord],
    fields_by_sheet: dict[str, list[FieldRecord]],
    confirmed_fields_by_sheet: dict[str, list[FieldRecord]],
    report: ValidationReport,
) -> tuple[list[CandidateRecord], dict[str, CandidateRecord], dict[int, CandidateRecord]]:
    records: list[CandidateRecord] = []
    by_id: dict[str, CandidateRecord] = {}
    by_product_id: dict[int, CandidateRecord] = {}
    confirmed_identity_keys: dict[tuple[str, str], str] = {}

    for item in rows:
        values = item.values
        candidate_id = clean_text(values["候选编号"])
        target_sheet = clean_text(values["目标参数表"])
        inclusion = clean_text(values["纳入状态"])
        identities = {field_name: values[field_name] for field_name in CANDIDATE_IDENTITY_FIELDS}
        raw_product_id = values["正式产品编号"]
        product_id = parse_positive_int(raw_product_id)

        if not CANDIDATE_ID_PATTERN.fullmatch(candidate_id):
            report.add(
                "CANDIDATE_ID_INVALID",
                "候选编号必须使用 CAND-0001 起的稳定文本格式。",
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=_column_cell(PRODUCT_CANDIDATE_HEADERS, "候选编号", item.row),
                candidate_id=candidate_id or None,
                evidence={"value": candidate_id},
            )
        if candidate_id in by_id:
            report.add(
                "CANDIDATE_ID_DUPLICATE",
                "候选编号不得重复。",
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        if not is_blank(raw_product_id) and product_id is None:
            code, message = _formal_product_id_error(raw_product_id, "正式产品编号")
            report.add(
                code,
                message,
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=_column_cell(PRODUCT_CANDIDATE_HEADERS, "正式产品编号", item.row),
                candidate_id=candidate_id or None,
                evidence={"value": raw_product_id, "type": type(raw_product_id).__name__},
            )
        if product_id is not None and product_id in by_product_id:
            report.add(
                "FORMAL_PRODUCT_ID_DUPLICATE",
                "同一正式产品编号不得分配给多个候选。",
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                product_id=product_id,
                evidence={"other_candidate": by_product_id[product_id].candidate_id},
            )
        if inclusion not in INCLUSION_STATES:
            report.add(
                "INCLUSION_STATE_INVALID",
                "纳入状态不在契约枚举中。",
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=_column_cell(PRODUCT_CANDIDATE_HEADERS, "纳入状态", item.row),
                candidate_id=candidate_id or None,
                product_id=product_id,
                evidence={"value": inclusion},
            )
        if not target_sheet or target_sheet not in fields_by_sheet:
            report.add(
                "TARGET_PARAMETER_SHEET_UNKNOWN",
                "目标参数表必须存在于字段设计中。",
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=_column_cell(PRODUCT_CANDIDATE_HEADERS, "目标参数表", item.row),
                candidate_id=candidate_id or None,
                product_id=product_id,
                evidence={"value": target_sheet},
            )
        if inclusion == "已确认纳入" and target_sheet not in confirmed_fields_by_sheet:
            report.add(
                "SAMPLE_CONFIRMATION_GATE_NOT_MET",
                "已确认纳入的产品只能指向已确认字段设计。",
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                product_id=product_id,
                evidence={"target_sheet": target_sheet},
            )
        if inclusion == "已确认纳入" and product_id is None:
            report.add(
                "CONFIRMED_PRODUCT_ID_MISSING",
                "已确认纳入的产品必须分配正式正整数编号。",
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        if inclusion in {"待核验", "待审核"} and product_id is not None:
            report.add(
                "PRODUCT_ID_ASSIGNED_BEFORE_CONFIRMATION",
                "用户确认纳入前不得分配正式产品编号。",
                workbook="process",
                sheet=PRODUCT_CANDIDATES_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                product_id=product_id,
            )
        for field_name, value in identities.items():
            if _forbidden_placeholder(value):
                report.add(
                    "PLACEHOLDER_NOT_ALLOWED",
                    "身份字段未知时必须留空，不得填写占位词。",
                    workbook="process",
                    sheet=PRODUCT_CANDIDATES_SHEET,
                    cell_or_row=_column_cell(PRODUCT_CANDIDATE_HEADERS, field_name, item.row),
                    candidate_id=candidate_id or None,
                    product_id=product_id,
                    evidence={"value": value},
                )
        if inclusion == "已确认纳入":
            for required_name in ("产品名称",):
                if is_blank(identities[required_name]):
                    report.add(
                        "CONFIRMED_PRODUCT_IDENTITY_MISSING",
                        "已确认产品必须具有产品名称。产品型号未由来源写明时保持空白。",
                        workbook="process",
                        sheet=PRODUCT_CANDIDATES_SHEET,
                        cell_or_row=_column_cell(PRODUCT_CANDIDATE_HEADERS, required_name, item.row),
                        candidate_id=candidate_id or None,
                        product_id=product_id,
                        evidence={"field": required_name},
                    )
            identity_key = (
                canonical_scalar(identities["产品名称"]),
                canonical_scalar(identities["产品型号"]),
            )
            if all(identity_key):
                if identity_key in confirmed_identity_keys:
                    report.add(
                        "PRODUCT_IDENTITY_DUPLICATE",
                        "同一产品名称和型号不得作为两个独立正式产品。",
                        workbook="process",
                        sheet=PRODUCT_CANDIDATES_SHEET,
                        cell_or_row=str(item.row),
                        candidate_id=candidate_id or None,
                        product_id=product_id,
                        evidence={"other_candidate": confirmed_identity_keys[identity_key]},
                    )
                confirmed_identity_keys.setdefault(identity_key, candidate_id)

        record = CandidateRecord(
            row=item.row,
            candidate_id=candidate_id,
            product_id=product_id,
            target_sheet=target_sheet,
            identities=identities,
            inclusion=inclusion,
        )
        records.append(record)
        if candidate_id and candidate_id not in by_id:
            by_id[candidate_id] = record
        if product_id is not None and product_id not in by_product_id:
            by_product_id[product_id] = record

    report.product_count = sum(record.inclusion == "已确认纳入" for record in records)
    return records, by_id, by_product_id


def _parse_parameter_evidence(
    rows: list[RowRecord],
    candidates_by_id: dict[str, CandidateRecord],
    fields_by_sheet: dict[str, list[FieldRecord]],
    confirmed_fields_by_sheet: dict[str, list[FieldRecord]],
    report: ValidationReport,
) -> tuple[list[ParameterEvidenceRecord], dict[tuple[str, str], ParameterEvidenceRecord]]:
    records: list[ParameterEvidenceRecord] = []
    evidence_ids: set[str] = set()
    main_evidence: dict[tuple[str, str], ParameterEvidenceRecord] = {}
    conflict_rows: dict[tuple[str, str], list[ParameterEvidenceRecord]] = defaultdict(list)

    for item in rows:
        values = item.values
        evidence_id = clean_text(values["证据编号"])
        candidate_id = clean_text(values["候选编号"])
        field_name = clean_text(values["字段名称"])
        source_type = clean_text(values["来源类型"])
        source_title = clean_text(values["来源标题"])
        source_url = clean_text(values["来源URL"])
        evidence_state = clean_text(values["证据状态"])
        adopted = clean_text(values["正式采用"])
        conflict_note = clean_text(values["冲突说明"])

        if not evidence_id:
            report.add(
                "EVIDENCE_ID_MISSING",
                "参数证据编号不能为空。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        elif evidence_id in evidence_ids:
            report.add(
                "EVIDENCE_ID_DUPLICATE",
                "参数证据编号不得重复。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                evidence={"evidence_id": evidence_id},
            )
        evidence_ids.add(evidence_id)

        candidate = candidates_by_id.get(candidate_id)
        if candidate is None:
            report.add(
                "EVIDENCE_CANDIDATE_ORPHAN",
                "参数证据引用了不存在的候选编号。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        if not field_name:
            report.add(
                "EVIDENCE_FIELD_MISSING",
                "参数证据字段名称不能为空。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        elif field_name == "产品编号":
            report.add(
                "INTERNAL_ID_EVIDENCE_FORBIDDEN",
                "内部生成的产品编号不得建立外部参数证据。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        elif candidate is not None:
            valid_fields = {record.field_name for record in fields_by_sheet.get(candidate.target_sheet, [])}
            if field_name not in valid_fields:
                report.add(
                    "EVIDENCE_FIELD_UNKNOWN",
                    "参数证据字段不属于该候选的目标参数表。",
                    workbook="process",
                    sheet=PARAMETER_EVIDENCE_SHEET,
                    cell_or_row=str(item.row),
                    candidate_id=candidate_id,
                    product_id=candidate.product_id,
                    evidence={"field": field_name, "target_sheet": candidate.target_sheet},
                )
        if source_type not in SOURCE_TYPES:
            report.add(
                "SOURCE_TYPE_INVALID",
                "参数来源类型不在契约枚举中。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                evidence={"value": source_type, "allowed": list(SOURCE_TYPES)},
            )
        if evidence_state not in EVIDENCE_STATES:
            report.add(
                "EVIDENCE_STATE_INVALID",
                "证据状态不在契约枚举中。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                evidence={"value": evidence_state},
            )
        if adopted not in ADOPTION_STATES:
            report.add(
                "ADOPTION_STATE_INVALID",
                "正式采用必须为“是”或“否”。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                evidence={"value": adopted},
            )
        if source_url and not is_http_url(source_url):
            report.add(
                "SOURCE_URL_INVALID",
                "参数来源 URL 必须是绝对 HTTP(S) 地址。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=_column_cell(PARAMETER_EVIDENCE_HEADERS, "来源URL", item.row),
                candidate_id=candidate_id or None,
                evidence={"value": source_url},
            )

        record = ParameterEvidenceRecord(
            row=item.row,
            evidence_id=evidence_id,
            candidate_id=candidate_id,
            field_name=field_name,
            original_value=values["原始值"],
            formal_value=values["正式值"],
            source_type=source_type,
            source_title=source_title,
            source_url=source_url,
            excerpt=clean_text(values["支持摘录"]),
            access_date=values["访问日期"],
            evidence_state=evidence_state,
            adopted=adopted,
            conflict_note=conflict_note,
        )
        records.append(record)
        key = (candidate_id, field_name)
        if evidence_state == "冲突":
            conflict_rows[key].append(record)

        is_main = evidence_state == "已确认" and adopted == "是"
        if adopted == "是" and evidence_state != "已确认":
            report.add(
                "EVIDENCE_ADOPTION_STATE_INVALID",
                "只有证据状态为“已确认”的参数证据可以正式采用。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        if is_main:
            if is_blank(record.formal_value):
                report.add(
                    "ADOPTED_FORMAL_VALUE_MISSING",
                    "正式采用的参数证据必须提供非空正式值。",
                    workbook="process",
                    sheet=PARAMETER_EVIDENCE_SHEET,
                    cell_or_row=str(item.row),
                    candidate_id=candidate_id or None,
                )
            if _forbidden_placeholder(record.formal_value):
                report.add(
                    "PLACEHOLDER_NOT_ALLOWED",
                    "正式值不得使用“未说明”等占位词；未知必须留空。",
                    workbook="process",
                    sheet=PARAMETER_EVIDENCE_SHEET,
                    cell_or_row=_column_cell(PARAMETER_EVIDENCE_HEADERS, "正式值", item.row),
                    candidate_id=candidate_id or None,
                    evidence={"value": record.formal_value},
                )
            if not source_url or not is_http_url(source_url):
                report.add(
                    "ADOPTED_EVIDENCE_SOURCE_REQUIRED",
                    "正式采用的参数事实必须指向绝对 HTTP(S) 来源。",
                    workbook="process",
                    sheet=PARAMETER_EVIDENCE_SHEET,
                    cell_or_row=str(item.row),
                    candidate_id=candidate_id or None,
                    evidence={"source_url": source_url},
                )
            if candidate is not None:
                designed_field = next(
                    (
                        field
                        for field in fields_by_sheet.get(candidate.target_sheet, [])
                        if field.field_name == field_name
                    ),
                    None,
                )
                if designed_field is not None:
                    try:
                        coerce_formal_value(record.formal_value, designed_field.value_type)
                    except ValueError as error:
                        report.add(
                            "FORMAL_VALUE_TYPE_INVALID",
                            "主证据正式值无法按字段设计的值类型写入正式工作簿。",
                            workbook="process",
                            sheet=PARAMETER_EVIDENCE_SHEET,
                            cell_or_row=_column_cell(PARAMETER_EVIDENCE_HEADERS, "正式值", item.row),
                            candidate_id=candidate_id,
                            product_id=candidate.product_id,
                            evidence={"field": field_name, "value_type": designed_field.value_type, "error": str(error)},
                        )
                confirmed_names = {
                    field.field_name
                    for field in confirmed_fields_by_sheet.get(candidate.target_sheet, [])
                }
                if field_name not in confirmed_names:
                    report.add(
                        "ADOPTED_FIELD_NOT_CONFIRMED",
                        "正式采用的参数证据只能对应已确认字段。",
                        workbook="process",
                        sheet=PARAMETER_EVIDENCE_SHEET,
                        cell_or_row=str(item.row),
                        candidate_id=candidate_id,
                        product_id=candidate.product_id,
                        evidence={"field": field_name},
                    )
                if candidate.inclusion in {"待核验", "待审核"}:
                    report.add(
                        "EVIDENCE_ADOPTED_BEFORE_PRODUCT_CONFIRMATION",
                        "产品确认纳入前不得标记正式采用证据。",
                        workbook="process",
                        sheet=PARAMETER_EVIDENCE_SHEET,
                        cell_or_row=str(item.row),
                        candidate_id=candidate_id,
                    )
            if key in main_evidence:
                report.add(
                    "MULTIPLE_ADOPTED_EVIDENCE",
                    "每个候选产品字段最多一条已确认且正式采用的主证据。",
                    workbook="process",
                    sheet=PARAMETER_EVIDENCE_SHEET,
                    cell_or_row=str(item.row),
                    candidate_id=candidate_id or None,
                    product_id=candidate.product_id if candidate else None,
                    evidence={"first_row": main_evidence[key].row, "second_row": item.row},
                )
            else:
                main_evidence[key] = record

    for key, conflicts in conflict_rows.items():
        if key in main_evidence and any(not item.conflict_note for item in conflicts):
            report.add(
                "UNRESOLVED_EVIDENCE_CONFLICT",
                "存在冲突证据时，冲突行必须记录解决说明后才能采用主证据。",
                workbook="process",
                sheet=PARAMETER_EVIDENCE_SHEET,
                cell_or_row=str(conflicts[0].row),
                candidate_id=key[0] or None,
                evidence={"field": key[1], "conflict_rows": [item.row for item in conflicts]},
            )

    for candidate in candidates_by_id.values():
        if candidate.inclusion != "已确认纳入":
            continue
        for field_name in CANDIDATE_IDENTITY_FIELDS:
            candidate_value = candidate.identities[field_name]
            main = main_evidence.get((candidate.candidate_id, field_name))
            if not is_blank(candidate_value) and main is None:
                report.add(
                    "IDENTITY_EVIDENCE_MISSING",
                    "每个非空正式身份值必须具有唯一主证据。",
                    workbook="process",
                    sheet=PRODUCT_CANDIDATES_SHEET,
                    cell_or_row=str(candidate.row),
                    candidate_id=candidate.candidate_id,
                    product_id=candidate.product_id,
                    evidence={"field": field_name, "value": candidate_value},
                )
            if main is not None and canonical_scalar(candidate_value) != canonical_scalar(main.formal_value):
                report.add(
                    "CANDIDATE_IDENTITY_EVIDENCE_MISMATCH",
                    "产品候选身份值必须与其主证据正式值一致。",
                    workbook="process",
                    sheet=PRODUCT_CANDIDATES_SHEET,
                    cell_or_row=str(candidate.row),
                    candidate_id=candidate.candidate_id,
                    product_id=candidate.product_id,
                    evidence={
                        "field": field_name,
                        "candidate_value": candidate_value,
                        "evidence_value": main.formal_value,
                    },
                )

    return records, main_evidence


def _parse_procurement_evidence(
    rows: list[RowRecord],
    candidates_by_id: dict[str, CandidateRecord],
    report: ValidationReport,
) -> tuple[list[ProcurementEvidenceRecord], list[ProcurementEvidenceRecord]]:
    records: list[ProcurementEvidenceRecord] = []
    adopted_records: list[ProcurementEvidenceRecord] = []
    channel_ids: set[str] = set()

    for item in rows:
        values = item.values
        channel_id = clean_text(values["渠道记录编号"])
        candidate_id = clean_text(values["候选编号"])
        evidence_state = clean_text(values["证据状态"])
        adopted = clean_text(values["正式采用"])
        business_values = {field_name: values[field_name] for field_name in PROCUREMENT_BUSINESS_FIELDS}
        candidate = candidates_by_id.get(candidate_id)

        if not channel_id:
            report.add(
                "CHANNEL_ID_MISSING",
                "渠道记录编号不能为空。",
                workbook="process",
                sheet=PROCUREMENT_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        elif channel_id in channel_ids:
            report.add(
                "CHANNEL_ID_DUPLICATE",
                "渠道记录编号不得重复。",
                workbook="process",
                sheet=PROCUREMENT_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                evidence={"channel_id": channel_id},
            )
        channel_ids.add(channel_id)
        if candidate is None:
            report.add(
                "PROCUREMENT_CANDIDATE_ORPHAN",
                "采购证据引用了不存在的候选编号。",
                workbook="process",
                sheet=PROCUREMENT_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        if evidence_state not in EVIDENCE_STATES:
            report.add(
                "PROCUREMENT_EVIDENCE_STATE_INVALID",
                "采购证据状态不在契约枚举中。",
                workbook="process",
                sheet=PROCUREMENT_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                evidence={"value": evidence_state},
            )
        if adopted not in ADOPTION_STATES:
            report.add(
                "PROCUREMENT_ADOPTION_STATE_INVALID",
                "采购记录正式采用必须为“是”或“否”。",
                workbook="process",
                sheet=PROCUREMENT_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
                evidence={"value": adopted},
            )

        record = ProcurementEvidenceRecord(
            row=item.row,
            channel_id=channel_id,
            candidate_id=candidate_id,
            business_values=business_values,
            source_title=clean_text(values["来源标题"]),
            excerpt=clean_text(values["支持摘录"]),
            access_date=values["访问日期"],
            evidence_state=evidence_state,
            adopted=adopted,
            conflict_note=clean_text(values["冲突说明"]),
        )
        records.append(record)

        for field_name in ("询价或购买链接", "国内代理/销售链接"):
            value = business_values[field_name]
            if not is_blank(value) and not is_http_url(value):
                report.add(
                    "PROCUREMENT_URL_INVALID",
                    "采购链接必须是绝对 HTTP(S) 地址。",
                    workbook="process",
                    sheet=PROCUREMENT_EVIDENCE_SHEET,
                    cell_or_row=_column_cell(PROCUREMENT_EVIDENCE_HEADERS, field_name, item.row),
                    candidate_id=candidate_id or None,
                    product_id=candidate.product_id if candidate else None,
                    evidence={"field": field_name, "value": value},
                )

        if adopted == "是" and evidence_state != "已确认":
            report.add(
                "PROCUREMENT_ADOPTION_STATE_MISMATCH",
                "只有已确认的采购证据可以正式采用。",
                workbook="process",
                sheet=PROCUREMENT_EVIDENCE_SHEET,
                cell_or_row=str(item.row),
                candidate_id=candidate_id or None,
            )
        if adopted == "是" and evidence_state == "已确认":
            adopted_records.append(record)
            if candidate is not None and candidate.inclusion in {"待核验", "待审核"}:
                report.add(
                    "PROCUREMENT_ADOPTED_BEFORE_PRODUCT_CONFIRMATION",
                    "产品确认纳入前不得正式采用采购记录。",
                    workbook="process",
                    sheet=PROCUREMENT_EVIDENCE_SHEET,
                    cell_or_row=str(item.row),
                    candidate_id=candidate_id,
                )
            for field_name, value in business_values.items():
                if _forbidden_placeholder(value):
                    report.add(
                        "PLACEHOLDER_NOT_ALLOWED",
                        "采购业务字段未知时必须留空，不得填写占位词。",
                        workbook="process",
                        sheet=PROCUREMENT_EVIDENCE_SHEET,
                        cell_or_row=_column_cell(PROCUREMENT_EVIDENCE_HEADERS, field_name, item.row),
                        candidate_id=candidate_id or None,
                        product_id=candidate.product_id if candidate else None,
                        evidence={"value": value},
                    )
            any_business_value = any(not is_blank(value) for value in business_values.values())
            if any_business_value and not (
                record.source_title or not is_blank(business_values["资料来源"])
            ):
                report.add(
                    "PROCUREMENT_SOURCE_REQUIRED",
                    "包含采购业务事实的正式采用记录必须保留采购来源。",
                    workbook="process",
                    sheet=PROCUREMENT_EVIDENCE_SHEET,
                    cell_or_row=str(item.row),
                    candidate_id=candidate_id or None,
                    product_id=candidate.product_id if candidate else None,
                )

    adopted_by_candidate = Counter(record.candidate_id for record in adopted_records)
    for candidate in candidates_by_id.values():
        if candidate.inclusion == "已确认纳入" and adopted_by_candidate[candidate.candidate_id] == 0:
            report.add(
                "PROCUREMENT_COVERAGE_MISSING",
                "每个正式产品至少需要一条已确认且正式采用的采购关联记录；无公开信息时业务字段留空。",
                workbook="process",
                sheet=PROCUREMENT_EVIDENCE_SHEET,
                candidate_id=candidate.candidate_id,
                product_id=candidate.product_id,
            )

    report.channel_count = sum(
        candidates_by_id.get(record.candidate_id) is not None
        and candidates_by_id[record.candidate_id].inclusion == "已确认纳入"
        for record in adopted_records
    )
    return records, adopted_records


def validate_process_workbook(
    process_path: str | Path, report: ValidationReport
) -> ProcessContext:
    workbook = load_workbook(process_path, data_only=False, read_only=False, keep_links=False)
    try:
        _validate_no_formulas(workbook, report, "process")
        parsed = _validate_process_sheet_structure(workbook, report)
        field_records, fields_by_sheet, confirmed_fields_by_sheet = _parse_fields(
            parsed.get(FIELD_DESIGN_SHEET, []), report
        )
        candidates, candidates_by_id, candidates_by_product_id = _parse_candidates(
            parsed.get(PRODUCT_CANDIDATES_SHEET, []),
            fields_by_sheet,
            confirmed_fields_by_sheet,
            report,
        )
        parameter_evidence, main_evidence = _parse_parameter_evidence(
            parsed.get(PARAMETER_EVIDENCE_SHEET, []),
            candidates_by_id,
            fields_by_sheet,
            confirmed_fields_by_sheet,
            report,
        )
        procurement_evidence, adopted_procurement = _parse_procurement_evidence(
            parsed.get(PROCUREMENT_EVIDENCE_SHEET, []), candidates_by_id, report
        )
    finally:
        workbook.close()

    return ProcessContext(
        field_records=field_records,
        fields_by_sheet=fields_by_sheet,
        confirmed_fields_by_sheet=confirmed_fields_by_sheet,
        candidates=candidates,
        candidates_by_id=candidates_by_id,
        candidates_by_product_id=candidates_by_product_id,
        parameter_evidence=parameter_evidence,
        main_evidence=main_evidence,
        procurement_evidence=procurement_evidence,
        adopted_procurement=adopted_procurement,
    )


def _xml_tag(namespace: str, local_name: str) -> str:
    return f"{{{namespace}}}{local_name}"


def _resolve_package_part(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def _read_relationships(archive: zipfile.ZipFile, part: str) -> dict[str, dict[str, str | None]]:
    rels_part = posixpath.join(
        posixpath.dirname(part), "_rels", posixpath.basename(part) + ".rels"
    )
    if rels_part not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read(rels_part))
    result: dict[str, dict[str, str | None]] = {}
    for relation in root.findall(_xml_tag(PACKAGE_REL_NS, "Relationship")):
        relation_id = relation.attrib.get("Id")
        if relation_id:
            result[relation_id] = {
                "type": relation.attrib.get("Type"),
                "target": relation.attrib.get("Target"),
                "target_mode": relation.attrib.get("TargetMode"),
            }
    return result


def _read_sheet_xml_info(path: str | Path) -> dict[str, SheetXmlInfo]:
    result: dict[str, SheetXmlInfo] = {}
    with zipfile.ZipFile(path, "r") as archive:
        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        workbook_relationships = _read_relationships(archive, "xl/workbook.xml")
        sheet_parts: dict[str, str] = {}
        for sheet in workbook_root.findall(f".//{_xml_tag(MAIN_NS, 'sheet')}"):
            name = sheet.attrib.get("name", "")
            relation_id = sheet.attrib.get(_xml_tag(OFFICE_REL_NS, "id"))
            relation = workbook_relationships.get(relation_id or "")
            target = relation.get("target") if relation else None
            if name and target:
                sheet_parts[name] = _resolve_package_part("xl/workbook.xml", target)

        for sheet_name, sheet_part in sheet_parts.items():
            if sheet_part not in archive.namelist():
                continue
            root = ET.fromstring(archive.read(sheet_part))
            info = SheetXmlInfo()
            worksheet_filter = root.find(_xml_tag(MAIN_NS, "autoFilter"))
            if worksheet_filter is not None:
                info.worksheet_auto_filter = worksheet_filter.attrib.get("ref")

            relationships = _read_relationships(archive, sheet_part)
            hyperlinks = root.find(_xml_tag(MAIN_NS, "hyperlinks"))
            if hyperlinks is not None:
                for hyperlink in hyperlinks.findall(_xml_tag(MAIN_NS, "hyperlink")):
                    relation_id = hyperlink.attrib.get(_xml_tag(OFFICE_REL_NS, "id"))
                    relation = relationships.get(relation_id or {})
                    info.hyperlinks.append(
                        {
                            "ref": hyperlink.attrib.get("ref"),
                            "relationship_id": relation_id,
                            "type": relation.get("type") if relation else None,
                            "target": relation.get("target") if relation else None,
                            "target_mode": relation.get("target_mode") if relation else None,
                        }
                    )

            table_parts = root.find(_xml_tag(MAIN_NS, "tableParts"))
            if table_parts is not None:
                for table_part in table_parts.findall(_xml_tag(MAIN_NS, "tablePart")):
                    relation_id = table_part.attrib.get(_xml_tag(OFFICE_REL_NS, "id"))
                    relation = relationships.get(relation_id or "")
                    target = relation.get("target") if relation else None
                    if not target:
                        continue
                    table_path = _resolve_package_part(sheet_part, target)
                    if table_path not in archive.namelist():
                        continue
                    table_root = ET.fromstring(archive.read(table_path))
                    table_ref = table_root.attrib.get("ref")
                    if table_ref:
                        info.table_refs.append(table_ref)
            result[sheet_name] = info
    return result


def _coordinate_in_ref(coordinate: str, reference: str | None) -> bool:
    if not reference:
        return False
    try:
        row, column = coordinate_to_tuple(coordinate)
        min_col, min_row, max_col, max_row = range_boundaries(reference)
    except ValueError:
        return False
    return min_row <= row <= max_row and min_col <= column <= max_col


def _hyperlink_for_coordinate(info: SheetXmlInfo, coordinate: str) -> dict[str, str | None] | None:
    for hyperlink in info.hyperlinks:
        if _coordinate_in_ref(coordinate, hyperlink.get("ref")):
            return hyperlink
    return None


def _formal_headers(worksheet: Any, row_number: int, width: int) -> tuple[str, ...]:
    return tuple(clean_text(worksheet.cell(row_number, column).value) for column in range(1, width + 1))


def _formal_data_rows(worksheet: Any, start_row: int) -> list[int]:
    result: list[int] = []
    for row_number in range(start_row, worksheet.max_row + 1):
        if any(not is_blank(cell.value) for cell in worksheet[row_number]):
            result.append(row_number)
    return result


def _identity_key(values: dict[str, Any]) -> tuple[str, str]:
    return (canonical_scalar(values.get("产品名称")), canonical_scalar(values.get("产品型号")))


def _formal_product_rows(
    process: ProcessContext,
    formal_path: str | Path,
    report: ValidationReport,
) -> tuple[dict[int, dict[str, Any]], dict[tuple[str, str], list[tuple[str, str, Any]]], list[tuple[str, str]]]:
    workbook = load_workbook(formal_path, data_only=False, read_only=False, keep_links=False)
    products_by_id: dict[int, dict[str, Any]] = {}
    formal_cells: dict[tuple[str, str], list[tuple[str, str, Any]]] = defaultdict(list)
    required_hyperlinks: list[tuple[str, str]] = []
    expected_sheet_names = set(process.confirmed_fields_by_sheet)

    try:
        _validate_no_formulas(workbook, report, "formal")

        expected_sheet_sequence = [
            sheet_name
            for sheet_name, _ in sorted(
                process.confirmed_fields_by_sheet.items(),
                key=lambda item: item[1][0].sheet_order,
            )
        ] + [PROCUREMENT_FORMAL_SHEET]
        if workbook.sheetnames != expected_sheet_sequence:
            report.add(
                "FORMAL_SHEET_ORDER_MISMATCH",
                "正式工作簿必须按字段设计中的工作表顺序排列参数表，并将采购渠道置于末尾。",
                workbook="formal",
                evidence={
                    "expected": expected_sheet_sequence,
                    "actual": workbook.sheetnames,
                },
            )

        for worksheet in workbook.worksheets:
            headers_to_scan = _formal_headers(worksheet, 1, worksheet.max_column)
            for column_number, header in enumerate(headers_to_scan, start=1):
                if header in FORBIDDEN_FORMAL_HEADERS:
                    report.add(
                        "FORMAL_PROCESS_FIELD_LEAK",
                        "正式工作簿不得出现研究过程字段。",
                        workbook="formal",
                        sheet=worksheet.title,
                        cell_or_row=f"{get_column_letter(column_number)}1",
                        evidence={"field": header},
                    )

            start_row = 2 if worksheet.title == PROCUREMENT_FORMAL_SHEET else 3
            for row_number in range(start_row, worksheet.max_row + 1):
                for cell in worksheet[row_number]:
                    if _forbidden_placeholder(cell.value):
                        report.add(
                            "PLACEHOLDER_NOT_ALLOWED",
                            "正式数据未知时必须留空，不得写入占位词。",
                            workbook="formal",
                            sheet=worksheet.title,
                            cell_or_row=cell.coordinate,
                            evidence={"value": cell.value},
                        )

        for sheet_name, fields in sorted(
            process.confirmed_fields_by_sheet.items(), key=lambda item: item[1][0].sheet_order
        ):
            if sheet_name not in workbook.sheetnames:
                report.add(
                    "FORMAL_PARAMETER_SHEET_MISSING",
                    "正式工作簿缺少已确认的参数表。",
                    workbook="formal",
                    sheet=sheet_name,
                )
                continue
            worksheet = workbook[sheet_name]
            expected_headers = tuple(field.field_name for field in fields)
            expected_units = tuple(field.unit for field in fields)
            actual_width = last_nonblank_column(
                worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)
            )
            if actual_width != len(expected_headers):
                report.add(
                    "FORMAL_PARAMETER_COLUMN_COUNT_INVALID",
                    "正式参数表不得增加、删除或隐藏已确认字段列。",
                    workbook="formal",
                    sheet=sheet_name,
                    cell_or_row="1",
                    evidence={"expected_count": len(expected_headers), "actual_count": actual_width},
                )
            actual_headers = _formal_headers(worksheet, 1, len(expected_headers))
            actual_units = _formal_headers(worksheet, 2, len(expected_headers))
            if actual_headers != expected_headers:
                report.add(
                    "FORMAL_HEADER_MISMATCH",
                    "正式参数表第一行字段名必须与已确认字段设计一致。",
                    workbook="formal",
                    sheet=sheet_name,
                    cell_or_row="1",
                    evidence={"expected": list(expected_headers), "actual": list(actual_headers)},
                )
            if actual_units != expected_units:
                report.add(
                    "FORMAL_UNIT_ROW_MISMATCH",
                    "正式参数表第二行必须是统一单位行；混合单位字段保持空白。",
                    workbook="formal",
                    sheet=sheet_name,
                    cell_or_row="2",
                    evidence={"expected": list(expected_units), "actual": list(actual_units)},
                )
            if worksheet.freeze_panes != "A3":
                report.add(
                    "FORMAL_FREEZE_PANE_INVALID",
                    "正式参数表必须冻结前两行（A3）。",
                    workbook="formal",
                    sheet=sheet_name,
                    evidence={"actual": str(worksheet.freeze_panes)},
                )
            if worksheet.auto_filter.ref:
                report.add(
                    "FORMAL_PARAMETER_AUTOFILTER_INVALID",
                    "参数表含两行表头，不得设置会把单位行当作数据的工作表级筛选。",
                    workbook="formal",
                    sheet=sheet_name,
                    evidence={"ref": worksheet.auto_filter.ref},
                )
            if list(worksheet.tables.values()):
                report.add(
                    "FORMAL_PARAMETER_TABLE_FORBIDDEN",
                    "正式参数表不创建 Excel Table，以避免单位行进入筛选数据。",
                    workbook="formal",
                    sheet=sheet_name,
                )

            header_to_field = {field.field_name: field for field in fields}
            for row_number in _formal_data_rows(worksheet, 3):
                row_values = {
                    field.field_name: worksheet.cell(row_number, index + 1).value
                    for index, field in enumerate(fields)
                }
                product_id = parse_positive_int(row_values["产品编号"])
                if product_id is None:
                    code, message = _formal_product_id_error(
                        row_values["产品编号"], "正式参数表产品编号"
                    )
                    report.add(
                        code,
                        message,
                        workbook="formal",
                        sheet=sheet_name,
                        cell_or_row=f"A{row_number}",
                        evidence={"value": row_values["产品编号"]},
                    )
                    continue
                if product_id in products_by_id:
                    report.add(
                        "FORMAL_PRODUCT_ID_DUPLICATE",
                        "一个正式产品只能在全部参数表中出现一次。",
                        workbook="formal",
                        sheet=sheet_name,
                        cell_or_row=f"A{row_number}",
                        product_id=product_id,
                        evidence={"other": products_by_id[product_id]["location"]},
                    )
                candidate = process.candidates_by_product_id.get(product_id)
                if candidate is None:
                    report.add(
                        "FORMAL_PRODUCT_ID_ORPHAN",
                        "正式参数表出现了过程账本中不存在的产品编号。",
                        workbook="formal",
                        sheet=sheet_name,
                        cell_or_row=f"A{row_number}",
                        product_id=product_id,
                    )
                    candidate_id = None
                else:
                    candidate_id = candidate.candidate_id
                    if candidate.inclusion != "已确认纳入":
                        report.add(
                            "FORMAL_PRODUCT_NOT_CONFIRMED",
                            "正式工作簿只能包含已确认纳入的产品。",
                            workbook="formal",
                            sheet=sheet_name,
                            cell_or_row=f"A{row_number}",
                            candidate_id=candidate_id,
                            product_id=product_id,
                        )
                    if candidate.target_sheet != sheet_name:
                        report.add(
                            "FORMAL_PRODUCT_CATEGORY_MISMATCH",
                            "产品只能出现在其过程账本指定的参数表中。",
                            workbook="formal",
                            sheet=sheet_name,
                            cell_or_row=f"A{row_number}",
                            candidate_id=candidate_id,
                            product_id=product_id,
                            evidence={"expected_sheet": candidate.target_sheet},
                        )

                products_by_id[product_id] = {
                    **row_values,
                    "candidate_id": candidate_id,
                    "sheet": sheet_name,
                    "row": row_number,
                    "location": f"{sheet_name}!A{row_number}",
                }
                if candidate_id:
                    for field in fields:
                        if field.field_name == "产品编号":
                            continue
                        cell = worksheet.cell(row_number, fields.index(field) + 1)
                        if not is_blank(cell.value):
                            formal_cells[(candidate_id, field.field_name)].append(
                                (sheet_name, cell.coordinate, cell.value)
                            )
                        if field.field_name in REQUIRED_HYPERLINK_FIELDS and not is_blank(cell.value):
                            required_hyperlinks.append((sheet_name, cell.coordinate))

        for worksheet in workbook.worksheets:
            if worksheet.title in expected_sheet_names or worksheet.title == PROCUREMENT_FORMAL_SHEET:
                continue
            report.add(
                "UNDECLARED_PARAMETER_SHEET",
                "正式工作簿不得包含未在确认字段设计中声明的工作表或默认产品总表。",
                workbook="formal",
                sheet=worksheet.title,
            )

        identity_keys: dict[tuple[str, str], int] = {}
        for product_id, values in products_by_id.items():
            key = _identity_key(values)
            if all(key):
                if key in identity_keys:
                    report.add(
                        "FORMAL_PRODUCT_IDENTITY_DUPLICATE",
                        "正式参数表不得重复同一产品名称和型号。",
                        workbook="formal",
                        sheet=values["sheet"],
                        cell_or_row=f"A{values['row']}",
                        product_id=product_id,
                        evidence={"other_product_id": identity_keys[key]},
                    )
                identity_keys.setdefault(key, product_id)

        if len(products_by_id) != report.product_count:
            report.add(
                "FORMAL_PRODUCT_COUNT_MISMATCH",
                "正式参数表产品数量必须等于已确认纳入产品数量。",
                workbook="formal",
                evidence={"formal": len(products_by_id), "process": report.product_count},
            )

        _validate_formal_parameter_evidence_mapping(process, products_by_id, formal_cells, report)
        _validate_formal_procurement(
            process, products_by_id, workbook, required_hyperlinks, report
        )
    finally:
        workbook.close()

    return products_by_id, formal_cells, required_hyperlinks


def _validate_formal_parameter_evidence_mapping(
    process: ProcessContext,
    products_by_id: dict[int, dict[str, Any]],
    formal_cells: dict[tuple[str, str], list[tuple[str, str, Any]]],
    report: ValidationReport,
) -> None:
    for product_id, values in products_by_id.items():
        candidate_id = values.get("candidate_id")
        if not candidate_id:
            continue
        candidate = process.candidates_by_id.get(candidate_id)
        if candidate is None:
            continue
        fields = process.confirmed_fields_by_sheet.get(candidate.target_sheet, [])
        for field in fields:
            if field.field_name == "产品编号":
                continue
            cells = formal_cells.get((candidate_id, field.field_name), [])
            main = process.main_evidence.get((candidate_id, field.field_name))
            if cells:
                if len(cells) != 1:
                    report.add(
                        "FORMAL_FIELD_MULTIPLE_CELLS",
                        "每个产品字段在正式参数表中只能有一个权威单元格。",
                        workbook="formal",
                        sheet=values["sheet"],
                        cell_or_row=str(values["row"]),
                        candidate_id=candidate_id,
                        product_id=product_id,
                        evidence={"field": field.field_name, "cells": cells},
                    )
                if main is None:
                    report.add(
                        "FORMAL_VALUE_EVIDENCE_MISSING",
                        "每个非空正式产品字段必须映射到唯一主证据。",
                        workbook="formal",
                        sheet=values["sheet"],
                        cell_or_row=cells[0][1],
                        candidate_id=candidate_id,
                        product_id=product_id,
                        evidence={"field": field.field_name, "value": cells[0][2]},
                    )
                elif canonical_scalar(cells[0][2]) != canonical_scalar(main.formal_value):
                    report.add(
                        "FORMAL_VALUE_EVIDENCE_MISMATCH",
                        "正式单元格值必须与主证据正式值一致。",
                        workbook="formal",
                        sheet=values["sheet"],
                        cell_or_row=cells[0][1],
                        candidate_id=candidate_id,
                        product_id=product_id,
                        evidence={
                            "field": field.field_name,
                            "formal_value": cells[0][2],
                            "evidence_value": main.formal_value,
                        },
                    )
            elif main is not None:
                report.add(
                    "ADOPTED_EVIDENCE_NOT_RENDERED",
                    "每条属于正式产品和已确认字段的主证据必须反向对应一个正式单元格。",
                    workbook="formal",
                    sheet=values["sheet"],
                    cell_or_row=str(values["row"]),
                    candidate_id=candidate_id,
                    product_id=product_id,
                    evidence={"field": field.field_name, "evidence_row": main.row},
                )


def _procurement_tuple(values: dict[str, Any]) -> tuple[str, ...]:
    return tuple(canonical_scalar(values.get(field_name)) for field_name in PROCUREMENT_BUSINESS_FIELDS)


def _validate_formal_procurement(
    process: ProcessContext,
    products_by_id: dict[int, dict[str, Any]],
    workbook: Any,
    required_hyperlinks: list[tuple[str, str]],
    report: ValidationReport,
) -> None:
    if PROCUREMENT_FORMAL_SHEET not in workbook.sheetnames:
        report.add(
            "FORMAL_PROCUREMENT_SHEET_MISSING",
            "正式工作簿必须包含固定 14 列的采购渠道表。",
            workbook="formal",
            sheet=PROCUREMENT_FORMAL_SHEET,
        )
        return

    worksheet = workbook[PROCUREMENT_FORMAL_SHEET]
    actual_width = last_nonblank_column(
        worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)
    )
    if actual_width != len(PROCUREMENT_FORMAL_HEADERS):
        report.add(
            "FORMAL_PROCUREMENT_COLUMN_COUNT_INVALID",
            "采购渠道表必须恰好包含固定 14 列。",
            workbook="formal",
            sheet=PROCUREMENT_FORMAL_SHEET,
            cell_or_row="1",
            evidence={"expected_count": len(PROCUREMENT_FORMAL_HEADERS), "actual_count": actual_width},
        )
    actual_headers = _formal_headers(worksheet, 1, len(PROCUREMENT_FORMAL_HEADERS))
    if actual_headers != PROCUREMENT_FORMAL_HEADERS:
        report.add(
            "FORMAL_PROCUREMENT_HEADER_MISMATCH",
            "采购渠道表必须使用固定 14 列单行表头。",
            workbook="formal",
            sheet=PROCUREMENT_FORMAL_SHEET,
            cell_or_row="1",
            evidence={"expected": list(PROCUREMENT_FORMAL_HEADERS), "actual": list(actual_headers)},
        )
    if worksheet.freeze_panes != "A2":
        report.add(
            "FORMAL_PROCUREMENT_FREEZE_INVALID",
            "采购渠道表必须冻结首行（A2）。",
            workbook="formal",
            sheet=PROCUREMENT_FORMAL_SHEET,
            evidence={"actual": str(worksheet.freeze_panes)},
        )
    tables = list(worksheet.tables.values())
    data_rows = _formal_data_rows(worksheet, 2)
    if len(tables) != 1:
        report.add(
            "FORMAL_PROCUREMENT_TABLE_INVALID",
            "采购渠道表必须且只能有一个 Excel Table。",
            workbook="formal",
            sheet=PROCUREMENT_FORMAL_SHEET,
            evidence={"table_count": len(tables)},
        )
    else:
        min_col, min_row, max_col, max_row = range_boundaries(tables[0].ref)
        if min_col != 1 or min_row != 1 or max_col != len(PROCUREMENT_FORMAL_HEADERS) or max_row < max(2, max(data_rows, default=2)):
            report.add(
                "FORMAL_PROCUREMENT_TABLE_RANGE_INVALID",
                "采购 Table 必须覆盖固定 14 列和全部渠道数据行。",
                workbook="formal",
                sheet=PROCUREMENT_FORMAL_SHEET,
                evidence={"table_ref": tables[0].ref, "data_rows": data_rows},
            )

    actual_counter: Counter[tuple[int, tuple[str, ...]]] = Counter()
    rows_by_product: Counter[int] = Counter()
    for row_number in data_rows:
        row_values = {
            field_name: worksheet.cell(row_number, index + 1).value
            for index, field_name in enumerate(PROCUREMENT_FORMAL_HEADERS)
        }
        product_id = parse_positive_int(row_values["产品编号"])
        if product_id is None:
            code, message = _formal_product_id_error(
                row_values["产品编号"], "采购渠道表产品编号"
            )
            report.add(
                code,
                message,
                workbook="formal",
                sheet=PROCUREMENT_FORMAL_SHEET,
                cell_or_row=f"A{row_number}",
                evidence={"value": row_values["产品编号"]},
            )
            continue
        rows_by_product[product_id] += 1
        if product_id not in products_by_id:
            report.add(
                "FORMAL_PROCUREMENT_PRODUCT_ORPHAN",
                "采购渠道表出现了不存在于参数表的产品编号。",
                workbook="formal",
                sheet=PROCUREMENT_FORMAL_SHEET,
                cell_or_row=f"A{row_number}",
                product_id=product_id,
            )
        else:
            product = products_by_id[product_id]
            for identity_field in ("产品名称", "产品型号"):
                if canonical_scalar(row_values[identity_field]) != canonical_scalar(product[identity_field]):
                    report.add(
                        "FORMAL_PROCUREMENT_IDENTITY_MISMATCH",
                        "采购渠道表的产品名称和型号必须与参数表一致。",
                        workbook="formal",
                        sheet=PROCUREMENT_FORMAL_SHEET,
                        cell_or_row=f"{get_column_letter(PROCUREMENT_FORMAL_HEADERS.index(identity_field)+1)}{row_number}",
                        product_id=product_id,
                        evidence={"field": identity_field, "expected": product[identity_field], "actual": row_values[identity_field]},
                    )
            actual_counter[(product_id, _procurement_tuple(row_values))] += 1
        for field_name in PROCUREMENT_BUSINESS_FIELDS:
            if _forbidden_placeholder(row_values[field_name]):
                report.add(
                    "PLACEHOLDER_NOT_ALLOWED",
                    "采购业务未知时必须留空，不得写入占位词。",
                    workbook="formal",
                    sheet=PROCUREMENT_FORMAL_SHEET,
                    cell_or_row=f"{get_column_letter(PROCUREMENT_FORMAL_HEADERS.index(field_name)+1)}{row_number}",
                    product_id=product_id,
                    evidence={"field": field_name, "value": row_values[field_name]},
                )
            if field_name in {"询价或购买链接", "国内代理/销售链接"} and not is_blank(row_values[field_name]):
                required_hyperlinks.append(
                    (PROCUREMENT_FORMAL_SHEET, f"{get_column_letter(PROCUREMENT_FORMAL_HEADERS.index(field_name)+1)}{row_number}")
                )
            if field_name == "资料来源" and is_http_url(row_values[field_name]):
                required_hyperlinks.append(
                    (PROCUREMENT_FORMAL_SHEET, f"{get_column_letter(PROCUREMENT_FORMAL_HEADERS.index(field_name)+1)}{row_number}")
                )

    expected_counter: Counter[tuple[int, tuple[str, ...]]] = Counter()
    for record in process.adopted_procurement:
        candidate = process.candidates_by_id.get(record.candidate_id)
        if candidate is not None and candidate.inclusion == "已确认纳入" and candidate.product_id is not None:
            expected_counter[(candidate.product_id, _procurement_tuple(record.business_values))] += 1

    if actual_counter != expected_counter:
        missing = list((expected_counter - actual_counter).elements())
        extra = list((actual_counter - expected_counter).elements())
        report.add(
            "FORMAL_PROCUREMENT_EVIDENCE_MISMATCH",
            "正式采购渠道行必须与已确认且正式采用的采购证据逐条对应。",
            workbook="formal",
            sheet=PROCUREMENT_FORMAL_SHEET,
            evidence={"missing": missing, "extra": extra},
        )
    for product_id in products_by_id:
        if rows_by_product[product_id] == 0:
            report.add(
                "FORMAL_PROCUREMENT_COVERAGE_MISSING",
                "每个正式产品至少需要一条采购渠道记录。",
                workbook="formal",
                sheet=PROCUREMENT_FORMAL_SHEET,
                product_id=product_id,
            )


def _read_previous_products(path: str | Path, report: ValidationReport) -> dict[int, tuple[str, str]]:
    workbook = load_workbook(path, data_only=False, read_only=False, keep_links=False)
    result: dict[int, tuple[str, str]] = {}
    try:
        for worksheet in workbook.worksheets:
            if worksheet.title == PROCUREMENT_FORMAL_SHEET:
                continue
            headers = _formal_headers(worksheet, 1, len(IDENTITY_FIELDS))
            if headers != IDENTITY_FIELDS:
                continue
            for row_number in _formal_data_rows(worksheet, 3):
                product_id = parse_positive_int(worksheet.cell(row_number, 1).value)
                if product_id is None:
                    report.add(
                        "PREVIOUS_FORMAL_PRODUCT_ID_INVALID",
                        "上一正式版本含非正整数产品编号。",
                        workbook="previous_formal",
                        sheet=worksheet.title,
                        cell_or_row=f"A{row_number}",
                        evidence={"value": worksheet.cell(row_number, 1).value},
                    )
                    continue
                identity = (
                    canonical_scalar(worksheet.cell(row_number, 3).value),
                    canonical_scalar(worksheet.cell(row_number, 4).value),
                )
                if product_id in result:
                    report.add(
                        "PREVIOUS_FORMAL_PRODUCT_ID_DUPLICATE",
                        "上一正式版本的产品编号不得重复。",
                        workbook="previous_formal",
                        sheet=worksheet.title,
                        cell_or_row=f"A{row_number}",
                        product_id=product_id,
                    )
                result[product_id] = identity
    finally:
        workbook.close()
    return result


def _validate_version_lifecycle(
    process: ProcessContext,
    products_by_id: dict[int, dict[str, Any]],
    previous_formal_path: str | Path | None,
    report: ValidationReport,
) -> None:
    current_ids = sorted(products_by_id)
    if previous_formal_path is None:
        expected = list(range(1, len(current_ids) + 1))
        if current_ids != expected:
            report.add(
                "INITIAL_PRODUCT_ID_SEQUENCE_INVALID",
                "首个正式版本的产品编号必须是全局连续的 1..N。",
                workbook="formal",
                evidence={"expected": expected, "actual": current_ids},
            )
        return

    previous = _read_previous_products(previous_formal_path, report)
    previous_max = max(previous, default=0)
    for product_id, identity in previous.items():
        candidate = process.candidates_by_product_id.get(product_id)
        if candidate is None:
            report.add(
                "PREVIOUS_ID_MISSING_FROM_LEDGER",
                "上一正式版本的产品编号必须在过程账本中保留，不能删除或清空。",
                workbook="process",
                candidate_id=None,
                product_id=product_id,
                evidence={"previous_identity": identity},
            )
            continue
        current_identity = _identity_key(candidate.identities)
        if current_identity != identity:
            report.add(
                "FORMAL_ID_REBOUND",
                "已发布产品编号不得改绑到其他产品名称或型号。",
                workbook="process",
                candidate_id=candidate.candidate_id,
                product_id=product_id,
                evidence={"previous_identity": identity, "current_identity": current_identity},
            )

    new_ids = []
    for product_id, values in products_by_id.items():
        identity = _identity_key(values)
        if product_id in previous:
            if identity != previous[product_id]:
                report.add(
                    "FORMAL_ID_REBOUND",
                    "正式版本更新不得改变旧编号对应的产品名称或型号。",
                    workbook="formal",
                    sheet=values["sheet"],
                    cell_or_row=f"A{values['row']}",
                    product_id=product_id,
                    evidence={"previous_identity": previous[product_id], "current_identity": identity},
                )
        else:
            new_ids.append(product_id)
            if product_id <= previous_max:
                report.add(
                    "FORMAL_ID_REUSED",
                    "新产品不得复用旧正式编号；编号必须大于上一版本最大编号。",
                    workbook="formal",
                    sheet=values["sheet"],
                    cell_or_row=f"A{values['row']}",
                    product_id=product_id,
                    evidence={"previous_max": previous_max},
                )
    expected_new_ids = list(range(previous_max + 1, previous_max + 1 + len(new_ids)))
    if sorted(new_ids) != expected_new_ids:
        report.add(
            "NEW_PRODUCT_ID_SEQUENCE_INVALID",
            "同一版本新增产品应从旧最大编号加一开始递增。",
            workbook="formal",
            evidence={"expected": expected_new_ids, "actual": sorted(new_ids)},
        )


def _valid_external_hyperlink(hyperlink: dict[str, str | None]) -> bool:
    relation_type = hyperlink.get("type") or ""
    target = hyperlink.get("target")
    return (
        relation_type.endswith("/hyperlink")
        and hyperlink.get("target_mode") == "External"
        and is_http_url(target)
    )


def _validate_ooxml_features(
    formal_path: str | Path,
    required_hyperlinks: list[tuple[str, str]],
    report: ValidationReport,
) -> None:
    infos = _read_sheet_xml_info(formal_path)
    for sheet_name, info in infos.items():
        for hyperlink in info.hyperlinks:
            if not _valid_external_hyperlink(hyperlink):
                report.add(
                    "HYPERLINK_RELATIONSHIP_INVALID",
                    "超链接必须具有可解析的 OOXML external hyperlink relationship 和绝对 HTTP(S) 目标。",
                    workbook="formal",
                    sheet=sheet_name,
                    cell_or_row=hyperlink.get("ref"),
                    evidence=hyperlink,
                )
        if info.worksheet_auto_filter:
            for table_ref in info.table_refs:
                try:
                    intersects = ranges_intersect(info.worksheet_auto_filter, table_ref)
                except ValueError:
                    intersects = True
                if intersects:
                    report.add(
                        "OVERLAPPING_AUTOFILTER",
                        "Excel Table 自带筛选，工作表级 autoFilter 不得与其范围相交。",
                        workbook="formal",
                        sheet=sheet_name,
                        evidence={
                            "worksheet_auto_filter": info.worksheet_auto_filter,
                            "table_ref": table_ref,
                        },
                    )

    for sheet_name, coordinate in sorted(set(required_hyperlinks)):
        info = infos.get(sheet_name)
        hyperlink = _hyperlink_for_coordinate(info, coordinate) if info else None
        if hyperlink is None:
            report.add(
                "HYPERLINK_RELATIONSHIP_MISSING",
                "非空产品页、规格书/手册或采购链接必须建立外部超链接关系。",
                workbook="formal",
                sheet=sheet_name,
                cell_or_row=coordinate,
            )
        elif not _valid_external_hyperlink(hyperlink):
            report.add(
                "HYPERLINK_RELATIONSHIP_INVALID",
                "链接关系必须指向绝对 HTTP(S) 外部目标。",
                workbook="formal",
                sheet=sheet_name,
                cell_or_row=coordinate,
                evidence=hyperlink,
            )


def validate_contract(
    process_path: str | Path,
    formal_path: str | Path | None = None,
    previous_formal_path: str | Path | None = None,
) -> ValidationReport:
    ensure_supported_openpyxl()
    if previous_formal_path is not None and formal_path is None:
        raise ValueError("--previous-formal requires --formal")
    inputs = {
        "process": _input_descriptor(process_path),
        "formal": _input_descriptor(formal_path),
        "previous_formal": _input_descriptor(previous_formal_path),
    }
    report = ValidationReport(
        mode="process_and_formal" if formal_path is not None else "process",
        inputs=inputs,
    )
    process = validate_process_workbook(process_path, report)
    if formal_path is not None:
        products_by_id, _, required_hyperlinks = _formal_product_rows(
            process, formal_path, report
        )
        _validate_ooxml_features(formal_path, required_hyperlinks, report)
        _validate_version_lifecycle(process, products_by_id, previous_formal_path, report)
        report.product_count = len(products_by_id)
    return report


def _unhashed_input_descriptor(path: str | Path | None) -> dict[str, str | None] | None:
    if path is None:
        return None
    return {"path": str(Path(path).absolute()), "sha256": None}


def _fatal_report(
    process_path: str | Path,
    formal_path: str | Path | None,
    previous_formal_path: str | Path | None,
    message: str,
) -> ValidationReport:
    report = ValidationReport(
        mode="process_and_formal" if formal_path is not None else "process",
        inputs={
            "process": _unhashed_input_descriptor(process_path),
            "formal": _unhashed_input_descriptor(formal_path),
            "previous_formal": _unhashed_input_descriptor(previous_formal_path),
        },
    )
    report.add("IO_OR_ARGUMENT_ERROR", message, workbook=None)
    return report


def _write_and_print(report: ValidationReport, json_out: str | None) -> int:
    payload = report.to_dict()
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    if json_out:
        write_json_no_overwrite(payload, json_out)
    return 0 if report.ok else 2


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Validate the process and formal product research workbook contract."
    )
    parser.add_argument("process", help="process workbook (.xlsx)")
    parser.add_argument("--formal", help="formal workbook (.xlsx)")
    parser.add_argument("--previous-formal", help="previous formal workbook (.xlsx)")
    parser.add_argument("--json-out", help="new JSON report path; existing paths are refused")
    try:
        arguments = parser.parse_args(argv)
    except SystemExit as error:
        return 0 if error.code == 0 else 1

    try:
        report = validate_contract(arguments.process, arguments.formal, arguments.previous_formal)
        return _write_and_print(report, arguments.json_out)
    except (OSError, ValueError, zipfile.BadZipFile, RuntimeError) as error:
        report = _fatal_report(
            arguments.process,
            arguments.formal,
            arguments.previous_formal,
            str(error),
        )
        try:
            _write_and_print(report, arguments.json_out)
        except (OSError, ValueError) as json_error:
            print(f"cannot write JSON report: {json_error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
