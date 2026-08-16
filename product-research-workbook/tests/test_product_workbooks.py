from __future__ import annotations

import io
import json
import sys
import tempfile
import unittest
from contextlib import redirect_stdout
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

SKILL_ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = SKILL_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from build_product_workbooks import build_formal, init_process
from inspect_product_samples import main as sample_inventory_main
from scan_legacy_identifiers import main as legacy_scan_main
from validate_product_workbooks import main as validator_main
from validate_product_workbooks import validate_contract
from workbook_contract import (
    FIELD_DESIGN_HEADERS,
    FIELD_DESIGN_SHEET,
    PARAMETER_EVIDENCE_HEADERS,
    PARAMETER_EVIDENCE_SHEET,
    PROCUREMENT_EVIDENCE_HEADERS,
    PROCUREMENT_EVIDENCE_SHEET,
    PROCUREMENT_FORMAL_HEADERS,
    PROCUREMENT_FORMAL_SHEET,
    PRODUCT_CANDIDATE_HEADERS,
    PRODUCT_CANDIDATES_SHEET,
    sha256_file,
)


def _column(headers: tuple[str, ...], header: str) -> int:
    return headers.index(header) + 1


def _row_by_value(worksheet, column: int, value: str) -> int:
    for row_number in range(2, worksheet.max_row + 1):
        if worksheet.cell(row_number, column).value == value:
            return row_number
    raise AssertionError(f"value not found: {value}")


def _refresh_table(worksheet, width: int) -> None:
    table = next(iter(worksheet.tables.values()))
    table.ref = f"A1:{chr(64 + width)}{max(2, worksheet.max_row)}"


def _append(worksheet, headers: tuple[str, ...], values: dict[str, object]) -> int:
    worksheet.append([values.get(header) for header in headers])
    _refresh_table(worksheet, len(headers))
    return worksheet.max_row


def _field_rows(sheet_order: int, sheet_name: str, technical_fields: list[tuple[str, str, str]]) -> list[dict[str, object]]:
    fields = [
        ("产品编号", "", "身份", "整数", ""),
        ("国家", "", "身份", "文本", ""),
        ("产品名称", "", "身份", "文本", ""),
        ("产品型号", "", "身份", "文本", ""),
        ("品牌", "", "身份", "文本", ""),
        ("生产厂家", "", "身份", "文本", ""),
        *[(name, unit, "技术参数", value_type, "真实产品样本和规格书交叉确认") for name, unit, value_type in technical_fields],
        ("产品页", "", "参数来源", "URL", ""),
        ("规格书/手册", "", "参数来源", "URL", ""),
        ("备注", "", "备注", "文本", ""),
    ]
    return [
        {
            "工作表顺序": sheet_order,
            "工作表名称": sheet_name,
            "字段顺序": order,
            "字段名称": name,
            "单位": unit,
            "字段角色": role,
            "值类型": value_type,
            "确认状态": "已确认",
            "样本依据": basis,
        }
        for order, (name, unit, role, value_type, basis) in enumerate(fields, start=1)
    ]


def _candidate_rows() -> list[dict[str, object]]:
    return [
        {
            "候选编号": "CAND-0001",
            "正式产品编号": 1,
            "目标参数表": "固定用户站",
            "国家": "中国",
            "产品名称": "固定用户地面站",
            "产品型号": None,
            "品牌": None,
            "生产厂家": None,
            "纳入状态": "已确认纳入",
            "复核备注": "",
        },
        {
            "候选编号": "CAND-0002",
            "正式产品编号": 2,
            "目标参数表": "固定用户站",
            "国家": "法国",
            "产品名称": "固定用户地面站",
            "产品型号": "GK-200",
            "品牌": "OrbitLink",
            "生产厂家": "OrbitLink SA",
            "纳入状态": "已确认纳入",
            "复核备注": "",
        },
        {
            "候选编号": "CAND-0003",
            "正式产品编号": 3,
            "目标参数表": "Ka相控阵组件",
            "国家": "美国",
            "产品名称": "Ka相控阵收发组件",
            "产品型号": "KA-01",
            "品牌": "ArrayWorks",
            "生产厂家": "ArrayWorks Inc.",
            "纳入状态": "已确认纳入",
            "复核备注": "",
        },
    ]


def _product_field_values(candidate: dict[str, object]) -> dict[str, object]:
    candidate_id = str(candidate["候选编号"])
    values = {
        "国家": candidate["国家"],
        "产品名称": candidate["产品名称"],
        "产品型号": candidate["产品型号"],
        "品牌": candidate["品牌"],
        "生产厂家": candidate["生产厂家"],
        "产品页": f"https://example.com/{candidate_id.lower()}",
        "规格书/手册": f"https://example.com/{candidate_id.lower()}/datasheet.pdf",
        "备注": "以官方资料限定条件为准",
    }
    if candidate["目标参数表"] == "固定用户站":
        values.update(
            {
                "接收频段": "17.7-20.2",
                "G/T": "18",
                "接口": "以太网",
                "典型功耗": "120",
                "峰值功耗": "180",
            }
        )
    else:
        values.update(
            {
                "工作频段": "17.7 GHz / 30 GHz",
                "接口": "LVDS",
                "峰值功耗": "65",
            }
        )
    return values


def _fill_valid_process(path: Path) -> None:
    with redirect_stdout(io.StringIO()):
        assert init_process(path) == 0
    workbook = load_workbook(path)
    try:
        for values in _field_rows(
            1,
            "固定用户站",
            [
                ("接收频段", "GHz", "文本"),
                ("G/T", "dB/K", "小数"),
                ("接口", "", "文本"),
                ("典型功耗", "W", "整数"),
                ("峰值功耗", "W", "整数"),
            ],
        ):
            _append(workbook[FIELD_DESIGN_SHEET], FIELD_DESIGN_HEADERS, values)
        for values in _field_rows(
            2,
            "Ka相控阵组件",
            [
                ("工作频段", "", "文本"),
                ("接口", "", "文本"),
                ("峰值功耗", "W", "整数"),
            ],
        ):
            _append(workbook[FIELD_DESIGN_SHEET], FIELD_DESIGN_HEADERS, values)

        candidates = _candidate_rows()
        for values in candidates:
            _append(workbook[PRODUCT_CANDIDATES_SHEET], PRODUCT_CANDIDATE_HEADERS, values)

        evidence_index = 1
        for candidate in candidates:
            for field_name, value in _product_field_values(candidate).items():
                if value is None:
                    continue
                source_type = "产品页" if field_name == "产品页" else "规格书/手册" if field_name == "规格书/手册" else "其他官方资料"
                _append(
                    workbook[PARAMETER_EVIDENCE_SHEET],
                    PARAMETER_EVIDENCE_HEADERS,
                    {
                        "证据编号": f"EVD-{evidence_index:04d}",
                        "候选编号": candidate["候选编号"],
                        "字段名称": field_name,
                        "原始值": value,
                        "正式值": value,
                        "来源类型": source_type,
                        "来源标题": "官方产品资料",
                        "来源URL": f"https://example.com/evidence/{evidence_index}",
                        "支持摘录": "官方页面或规格书中的对应参数。",
                        "页码/章节": "规格参数",
                        "访问日期": date(2026, 8, 16),
                        "证据状态": "已确认",
                        "正式采用": "是",
                        "冲突说明": "",
                    },
                )
                evidence_index += 1

        procurement_rows = [
            {
                "渠道记录编号": "CH-0001",
                "候选编号": "CAND-0001",
                "渠道公司": "示例渠道公司",
                "渠道角色": "制造商",
                "可销售地区": "中国",
                "公开价格": None,
                "最小起订量": None,
                "电话": None,
                "邮箱": "sales@example.com",
                "询价或购买链接": "https://example.com/buy/gk-100",
                "国内代理/销售链接": None,
                "采购或供货说明": "通过公开询价入口联系。",
                "资料来源": "https://example.com/channel/gk-100",
                "来源标题": "制造商公开页面",
            },
            {
                "渠道记录编号": "CH-0002",
                "候选编号": "CAND-0002",
                "渠道公司": "Europe Channel",
                "渠道角色": "授权渠道",
                "可销售地区": "欧洲",
                "公开价格": None,
                "最小起订量": None,
                "电话": None,
                "邮箱": None,
                "询价或购买链接": "https://example.com/buy/gk-200",
                "国内代理/销售链接": None,
                "采购或供货说明": "公开询价。",
                "资料来源": "渠道公开页面",
                "来源标题": "授权渠道页面",
            },
            {
                "渠道记录编号": "CH-0003",
                "候选编号": "CAND-0003",
                "渠道公司": None,
                "渠道角色": None,
                "可销售地区": None,
                "公开价格": None,
                "最小起订量": None,
                "电话": None,
                "邮箱": None,
                "询价或购买链接": None,
                "国内代理/销售链接": None,
                "采购或供货说明": None,
                "资料来源": None,
                "来源标题": None,
            },
        ]
        for values in procurement_rows:
            values.update(
                {
                    "支持摘录": "采购来源记录。",
                    "访问日期": date(2026, 8, 16),
                    "证据状态": "已确认",
                    "正式采用": "是",
                    "冲突说明": "",
                }
            )
            _append(workbook[PROCUREMENT_EVIDENCE_SHEET], PROCUREMENT_EVIDENCE_HEADERS, values)
        workbook.save(path)
    finally:
        workbook.close()


class ProductWorkbookContractTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.process = self.root / "process.xlsx"
        _fill_valid_process(self.process)

    def tearDown(self) -> None:
        self.temporary_directory.cleanup()

    def _build(self, name: str = "formal.xlsx", previous: Path | None = None) -> Path:
        output = self.root / name
        with redirect_stdout(io.StringIO()):
            result = build_formal(self.process, output, previous)
        self.assertEqual(result, 0)
        self.assertTrue(output.exists())
        return output

    def _codes(self, report) -> set[str]:
        return {finding.code for finding in report.findings}

    def test_valid_two_categories_blank_optional_identity_and_empty_channel_pass(self) -> None:
        formal = self._build()
        report = validate_contract(self.process, formal)
        self.assertTrue(report.ok, report.to_dict())
        self.assertEqual(report.product_count, 3)
        self.assertEqual(report.channel_count, 3)
        workbook = load_workbook(formal, data_only=False)
        try:
            self.assertEqual(workbook.sheetnames, ["固定用户站", "Ka相控阵组件", PROCUREMENT_FORMAL_SHEET])
            self.assertIsNone(workbook["固定用户站"].cell(3, 4).value)
            self.assertIsNone(workbook["固定用户站"].cell(3, 5).value)
            self.assertEqual(workbook["Ka相控阵组件"].cell(2, 7).value, None)
            self.assertIsNone(workbook[PROCUREMENT_FORMAL_SHEET].cell(2, 3).value)
            self.assertEqual(workbook[PROCUREMENT_FORMAL_SHEET].max_row, 4)
            self.assertEqual(sum(1 for ws in workbook.worksheets for row in ws.iter_rows() for cell in row if cell.data_type == "f"), 0)
        finally:
            workbook.close()

    def test_whole_machine_sheet_is_first_and_formal_order_is_validated(self) -> None:
        whole_machine_sheet = "固定用户站整机"
        workbook = load_workbook(self.process)
        try:
            fields = workbook[FIELD_DESIGN_SHEET]
            field_sheet_column = _column(FIELD_DESIGN_HEADERS, "工作表名称")
            for row_number in range(2, fields.max_row + 1):
                if fields.cell(row_number, field_sheet_column).value == "固定用户站":
                    fields.cell(row_number, field_sheet_column, whole_machine_sheet)

            candidates = workbook[PRODUCT_CANDIDATES_SHEET]
            target_sheet_column = _column(PRODUCT_CANDIDATE_HEADERS, "目标参数表")
            for row_number in range(2, candidates.max_row + 1):
                if candidates.cell(row_number, target_sheet_column).value == "固定用户站":
                    candidates.cell(row_number, target_sheet_column, whole_machine_sheet)
            workbook.save(self.process)
        finally:
            workbook.close()

        formal = self._build("whole-machine-first.xlsx")
        workbook = load_workbook(formal)
        try:
            self.assertEqual(
                workbook.sheetnames,
                [whole_machine_sheet, "Ka相控阵组件", PROCUREMENT_FORMAL_SHEET],
            )
        finally:
            workbook.close()

        reordered = self.root / "whole-machine-second.xlsx"
        workbook = load_workbook(formal)
        try:
            workbook._sheets = [
                workbook["Ka相控阵组件"],
                workbook[whole_machine_sheet],
                workbook[PROCUREMENT_FORMAL_SHEET],
            ]
            workbook.save(reordered)
        finally:
            workbook.close()

        report = validate_contract(self.process, reordered)
        self.assertIn("FORMAL_SHEET_ORDER_MISMATCH", self._codes(report))

    def test_textual_legacy_identifier_is_rejected_before_generation(self) -> None:
        workbook = load_workbook(self.process)
        try:
            candidates = workbook[PRODUCT_CANDIDATES_SHEET]
            row = _row_by_value(candidates, _column(PRODUCT_CANDIDATE_HEADERS, "候选编号"), "CAND-0001")
            candidates.cell(row, _column(PRODUCT_CANDIDATE_HEADERS, "正式产品编号"), "LEGACY-ROW-001")
            workbook.save(self.process)
        finally:
            workbook.close()

        report = validate_contract(self.process)
        self.assertIn("FORMAL_PRODUCT_ID_TEXT_FORBIDDEN", self._codes(report))
        output = self.root / "legacy-label-formal.xlsx"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(build_formal(self.process, output, None), 2)
        self.assertFalse(output.exists())

    def test_noncanonical_candidate_identifier_is_rejected_before_generation(self) -> None:
        workbook = load_workbook(self.process)
        try:
            candidates = workbook[PRODUCT_CANDIDATES_SHEET]
            row = _row_by_value(candidates, _column(PRODUCT_CANDIDATE_HEADERS, "候选编号"), "CAND-0001")
            candidates.cell(row, _column(PRODUCT_CANDIDATE_HEADERS, "候选编号"), "ARCHIVE-ROW-0001")
            workbook.save(self.process)
        finally:
            workbook.close()

        report = validate_contract(self.process)
        self.assertIn("CANDIDATE_ID_INVALID", self._codes(report))
        output = self.root / "invalid-candidate-formal.xlsx"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(build_formal(self.process, output, None), 2)
        self.assertFalse(output.exists())

    def test_legacy_identifier_scan_detects_leak_without_echoing_identifier(self) -> None:
        source = self.root / "legacy-source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "历史产品"
        worksheet.append(["产品编号", "产品名称"])
        worksheet.append(["ARCHIVE-ROW-0001", "测试产品"])
        workbook.save(source)
        workbook.close()

        output_directory = self.root / "staging"
        output_directory.mkdir()
        (output_directory / "report.md").write_text("ARCHIVE-ROW-0001", encoding="utf-8")
        failure_report = self.root / "legacy-scan-failure.json"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(
                legacy_scan_main(
                    [str(source), str(output_directory), "--json-out", str(failure_report)]
                ),
                2,
            )
        failure_payload = json.loads(failure_report.read_text(encoding="utf-8"))
        self.assertFalse(failure_payload["ok"])
        self.assertEqual(failure_payload["findings"][0]["artifact"], "report.md")
        self.assertNotIn("ARCHIVE-ROW-0001", json.dumps(failure_payload, ensure_ascii=False))

        (output_directory / "report.md").unlink()
        nested_output = output_directory / "ARCHIVE-ROW-0001"
        nested_output.mkdir()
        (nested_output / "safe.md").write_text("源表第 2 行", encoding="utf-8")
        nested_failure_report = self.root / "legacy-scan-nested-failure.json"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(
                legacy_scan_main(
                    [str(source), str(output_directory), "--json-out", str(nested_failure_report)]
                ),
                2,
            )
        nested_failure_payload = json.loads(nested_failure_report.read_text(encoding="utf-8"))
        self.assertFalse(nested_failure_payload["ok"])
        self.assertEqual(nested_failure_payload["findings"][0]["artifact"], "<redacted-path>")
        self.assertNotIn("ARCHIVE-ROW-0001", json.dumps(nested_failure_payload, ensure_ascii=False))
        (nested_output / "safe.md").unlink()
        nested_output.rmdir()

        (output_directory / "safe.md").write_text("源表第 2 行", encoding="utf-8")
        success_report = self.root / "legacy-scan-success.json"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(
                legacy_scan_main(
                    [str(source), str(output_directory), "--json-out", str(success_report)]
                ),
                0,
            )
        self.assertTrue(json.loads(success_report.read_text(encoding="utf-8"))["ok"])

    def test_sanitized_sample_inventory_excludes_identifier_values(self) -> None:
        source = self.root / "legacy-source.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "历史产品"
        worksheet.append(["产品编号", "产品名称", "产品型号"])
        worksheet.append(["ARCHIVE-ROW-0001", "测试产品", "MODEL-01"])
        worksheet.append(["ARCHIVE-ROW-0002", "测试组件", "MODULE-02"])
        workbook.save(source)
        workbook.close()

        staging = self.root / "sanitized-staging"
        staging.mkdir()
        inventory = staging / "sample-inventory.json"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(
                sample_inventory_main(
                    [str(source), str(inventory), "--identifier-header", "产品编号", "--sample-rows", "2"]
                ),
                0,
            )
        payload = json.loads(inventory.read_text(encoding="utf-8"))
        self.assertEqual(payload["identifier_column_count"], 1)
        section = payload["worksheets"][0]["sections"][0]
        self.assertNotIn("产品编号", section["headers"])
        serialized = json.dumps(payload, ensure_ascii=False)
        self.assertNotIn("ARCHIVE-ROW-0001", serialized)
        self.assertNotIn("ARCHIVE-ROW-0002", serialized)

        scan_report = self.root / "sanitized-scan.json"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(
                legacy_scan_main([str(source), str(staging), "--json-out", str(scan_report)]),
                0,
            )

    def test_textual_legacy_identifier_is_rejected_from_formal_output(self) -> None:
        formal = self._build()
        workbook = load_workbook(formal)
        try:
            workbook["固定用户站"].cell(3, 1, "LEGACY-ROW-001")
            workbook[PROCUREMENT_FORMAL_SHEET].cell(2, 1, "LEGACY-ROW-001")
            workbook.save(formal)
        finally:
            workbook.close()

        findings = [
            finding
            for finding in validate_contract(self.process, formal).findings
            if finding.code == "FORMAL_PRODUCT_ID_TEXT_FORBIDDEN"
        ]
        locations = {(finding.sheet, finding.cell_or_row) for finding in findings}
        self.assertIn(("固定用户站", "A3"), locations)
        self.assertIn((PROCUREMENT_FORMAL_SHEET, "A2"), locations)

    def test_undeclared_default_product_summary_sheet_is_rejected(self) -> None:
        formal = self._build()
        workbook = load_workbook(formal)
        try:
            workbook.create_sheet("产品总表")
            workbook.save(formal)
        finally:
            workbook.close()

        findings = [
            finding
            for finding in validate_contract(self.process, formal).findings
            if finding.code == "UNDECLARED_PARAMETER_SHEET"
        ]
        self.assertEqual([finding.sheet for finding in findings], ["产品总表"])

    def test_duplicate_identifier_orphan_and_missing_procurement_fail(self) -> None:
        workbook = load_workbook(self.process)
        try:
            candidates = workbook[PRODUCT_CANDIDATES_SHEET]
            row = _row_by_value(candidates, _column(PRODUCT_CANDIDATE_HEADERS, "候选编号"), "CAND-0002")
            candidates.cell(row, _column(PRODUCT_CANDIDATE_HEADERS, "正式产品编号"), 1)
            procurement = workbook[PROCUREMENT_EVIDENCE_SHEET]
            row = _row_by_value(procurement, _column(PROCUREMENT_EVIDENCE_HEADERS, "候选编号"), "CAND-0003")
            procurement.delete_rows(row)
            _refresh_table(procurement, len(PROCUREMENT_EVIDENCE_HEADERS))
            workbook.save(self.process)
        finally:
            workbook.close()
        report = validate_contract(self.process)
        codes = self._codes(report)
        self.assertIn("FORMAL_PRODUCT_ID_DUPLICATE", codes)
        self.assertIn("PROCUREMENT_COVERAGE_MISSING", codes)

        _fill_valid_process(self.process.with_name("replacement.xlsx"))
        replacement = self.process.with_name("replacement.xlsx")
        formal = self.root / "orphan.xlsx"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(build_formal(replacement, formal, None), 0)
        workbook = load_workbook(formal)
        try:
            workbook[PROCUREMENT_FORMAL_SHEET].cell(2, 1, 999)
            workbook.save(formal)
        finally:
            workbook.close()
        report = validate_contract(replacement, formal)
        self.assertIn("FORMAL_PROCUREMENT_PRODUCT_ORPHAN", self._codes(report))

    def test_unit_row_process_field_and_placeholder_fail(self) -> None:
        formal = self._build()
        workbook = load_workbook(formal)
        try:
            component = workbook["Ka相控阵组件"]
            component.cell(2, 7, "GHz")
            component.cell(1, 7, "证据状态")
            component.cell(3, 8, "未说明")
            workbook.save(formal)
        finally:
            workbook.close()
        codes = self._codes(validate_contract(self.process, formal))
        self.assertIn("FORMAL_UNIT_ROW_MISMATCH", codes)
        self.assertIn("FORMAL_PROCESS_FIELD_LEAK", codes)
        self.assertIn("PLACEHOLDER_NOT_ALLOWED", codes)

    def test_multiple_adoption_and_unmapped_evidence_fail(self) -> None:
        workbook = load_workbook(self.process)
        try:
            evidence = workbook[PARAMETER_EVIDENCE_SHEET]
            source_row = _row_by_value(evidence, _column(PARAMETER_EVIDENCE_HEADERS, "字段名称"), "接收频段")
            values = {
                header: evidence.cell(source_row, index + 1).value
                for index, header in enumerate(PARAMETER_EVIDENCE_HEADERS)
            }
            values["证据编号"] = "EVD-9999"
            _append(evidence, PARAMETER_EVIDENCE_HEADERS, values)
            workbook.save(self.process)
        finally:
            workbook.close()
        self.assertIn("MULTIPLE_ADOPTED_EVIDENCE", self._codes(validate_contract(self.process)))

        _fill_valid_process(self.process.with_name("mapping.xlsx"))
        mapping_process = self.process.with_name("mapping.xlsx")
        mapping_formal = self.root / "mapping-formal.xlsx"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(build_formal(mapping_process, mapping_formal, None), 0)
        workbook = load_workbook(mapping_formal)
        try:
            workbook["固定用户站"].cell(3, 7).value = None
            workbook.save(mapping_formal)
        finally:
            workbook.close()
        self.assertIn(
            "ADOPTED_EVIDENCE_NOT_RENDERED",
            self._codes(validate_contract(mapping_process, mapping_formal)),
        )

    def test_hyperlink_and_filter_overlap_fail(self) -> None:
        formal = self._build()
        workbook = load_workbook(formal)
        try:
            workbook["固定用户站"].cell(3, 12).hyperlink = None
            workbook["Ka相控阵组件"].cell(3, 10).hyperlink = "ftp://example.com/not-allowed"
            procurement = workbook[PROCUREMENT_FORMAL_SHEET]
            procurement.auto_filter.ref = next(iter(procurement.tables.values())).ref
            workbook.save(formal)
        finally:
            workbook.close()
        codes = self._codes(validate_contract(self.process, formal))
        self.assertIn("HYPERLINK_RELATIONSHIP_MISSING", codes)
        self.assertIn("HYPERLINK_RELATIONSHIP_INVALID", codes)
        self.assertIn("OVERLAPPING_AUTOFILTER", codes)

    def test_output_overwrite_is_refused_and_cli_report_is_stable(self) -> None:
        with self.assertRaises(FileExistsError):
            init_process(self.process)
        existing = self.root / "existing.xlsx"
        existing.write_bytes(b"not a workbook")
        with self.assertRaises(FileExistsError):
            build_formal(self.process, existing, None)

        report_path = self.root / "report.json"
        with redirect_stdout(io.StringIO()):
            self.assertEqual(validator_main([str(self.process), "--json-out", str(report_path)]), 0)
        self.assertTrue(report_path.exists())
        with redirect_stdout(io.StringIO()):
            self.assertEqual(validator_main([str(self.process), "--json-out", str(report_path)]), 1)

    def test_build_report_points_to_published_formal_path(self) -> None:
        output = self.root / "reported-formal.xlsx"
        stdout = io.StringIO()
        with redirect_stdout(stdout):
            self.assertEqual(build_formal(self.process, output, None), 0)
        payload = json.loads(stdout.getvalue())
        self.assertEqual(payload["output"]["path"], str(output.resolve()))
        self.assertEqual(payload["output"]["sha256"], sha256_file(output))
        self.assertEqual(payload["validation"]["inputs"]["formal"]["path"], str(output.resolve()))
        self.assertEqual(payload["validation"]["inputs"]["formal"]["sha256"], payload["output"]["sha256"])

    def test_cross_version_id_stability_and_no_reuse(self) -> None:
        version_one = self._build("formal-v1.xlsx")
        workbook = load_workbook(self.process)
        try:
            candidates = workbook[PRODUCT_CANDIDATES_SHEET]
            row = _row_by_value(candidates, _column(PRODUCT_CANDIDATE_HEADERS, "候选编号"), "CAND-0002")
            candidates.cell(row, _column(PRODUCT_CANDIDATE_HEADERS, "纳入状态"), "已排除")
            new_candidate = {
                "候选编号": "CAND-0004",
                "正式产品编号": 4,
                "目标参数表": "固定用户站",
                "国家": "日本",
                "产品名称": "固定用户地面站",
                "产品型号": "GK-400",
                "品牌": "SkyConnect",
                "生产厂家": "SkyConnect KK",
                "纳入状态": "已确认纳入",
                "复核备注": "",
            }
            _append(candidates, PRODUCT_CANDIDATE_HEADERS, new_candidate)
            evidence = workbook[PARAMETER_EVIDENCE_SHEET]
            next_evidence = evidence.max_row + 1
            for offset, (field_name, value) in enumerate(_product_field_values(new_candidate).items(), start=0):
                if value is None:
                    continue
                _append(
                    evidence,
                    PARAMETER_EVIDENCE_HEADERS,
                    {
                        "证据编号": f"EVD-{next_evidence + offset:04d}",
                        "候选编号": "CAND-0004",
                        "字段名称": field_name,
                        "原始值": value,
                        "正式值": value,
                        "来源类型": "产品页" if field_name == "产品页" else "规格书/手册" if field_name == "规格书/手册" else "其他官方资料",
                        "来源标题": "官方资料",
                        "来源URL": f"https://example.com/new/{next_evidence + offset}",
                        "支持摘录": "证据。",
                        "页码/章节": "规格",
                        "访问日期": date(2026, 8, 16),
                        "证据状态": "已确认",
                        "正式采用": "是",
                        "冲突说明": "",
                    },
                )
            _append(
                workbook[PROCUREMENT_EVIDENCE_SHEET],
                PROCUREMENT_EVIDENCE_HEADERS,
                {
                    "渠道记录编号": "CH-0004",
                    "候选编号": "CAND-0004",
                    "渠道公司": None,
                    "渠道角色": None,
                    "可销售地区": None,
                    "公开价格": None,
                    "最小起订量": None,
                    "电话": None,
                    "邮箱": None,
                    "询价或购买链接": None,
                    "国内代理/销售链接": None,
                    "采购或供货说明": None,
                    "资料来源": None,
                    "来源标题": None,
                    "支持摘录": "",
                    "访问日期": date(2026, 8, 16),
                    "证据状态": "已确认",
                    "正式采用": "是",
                    "冲突说明": "",
                },
            )
            workbook.save(self.process)
        finally:
            workbook.close()

        version_two = self._build("formal-v2.xlsx", version_one)
        report = validate_contract(self.process, version_two, version_one)
        self.assertTrue(report.ok, report.to_dict())
        workbook = load_workbook(version_two, read_only=True)
        try:
            ids = []
            for sheet_name in ("固定用户站", "Ka相控阵组件"):
                ids.extend(workbook[sheet_name].cell(row, 1).value for row in range(3, workbook[sheet_name].max_row + 1))
            self.assertEqual(sorted(ids), [1, 3, 4])
        finally:
            workbook.close()

        workbook = load_workbook(self.process)
        try:
            candidates = workbook[PRODUCT_CANDIDATES_SHEET]
            row = _row_by_value(candidates, _column(PRODUCT_CANDIDATE_HEADERS, "候选编号"), "CAND-0002")
            candidates.cell(row, _column(PRODUCT_CANDIDATE_HEADERS, "纳入状态"), "已确认纳入")
            workbook.save(self.process)
        finally:
            workbook.close()
        output = io.StringIO()
        with redirect_stdout(output):
            self.assertEqual(build_formal(self.process, self.root / "formal-v3.xlsx", version_two), 2)
        payload = json.loads(output.getvalue())
        self.assertIn("FORMAL_ID_REUSED", {item["code"] for item in payload["findings"]})


if __name__ == "__main__":
    unittest.main()
