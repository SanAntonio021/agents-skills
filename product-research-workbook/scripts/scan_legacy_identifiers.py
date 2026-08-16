from __future__ import annotations

import argparse
import json
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from workbook_contract import CANDIDATE_ID_PATTERN, clean_text, sha256_file, write_json_no_overwrite


SCAN_SCHEMA_VERSION = "1.0"
DEFAULT_IDENTIFIER_HEADERS = ("产品编号",)


@dataclass(frozen=True)
class ScanFinding:
    artifact: str
    location_type: str
    match_count: int

    def to_dict(self) -> dict[str, Any]:
        return {
            "code": "LEGACY_IDENTIFIER_LEAK",
            "severity": "error",
            "artifact": self.artifact,
            "location_type": self.location_type,
            "match_count": self.match_count,
        }


def _looks_like_manual_identifier(value: Any) -> bool:
    text = clean_text(value)
    return (
        bool(text)
        and not text.isdecimal()
        and not CANDIDATE_ID_PATTERN.fullmatch(text)
        and any(character.isdigit() for character in text)
        and ("-" in text or "_" in text)
    )


def _extract_legacy_identifiers(source_path: Path, headers: tuple[str, ...]) -> tuple[set[str], int]:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    identifiers: set[str] = set()
    matched_header_count = 0
    try:
        wanted_headers = {header.strip() for header in headers if header.strip()}
        for worksheet in workbook.worksheets:
            header_cells: list[tuple[int, int]] = []
            for row in worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row)):
                for cell in row:
                    if clean_text(cell.value) in wanted_headers:
                        header_cells.append((cell.row, cell.column))
            matched_header_count += len(header_cells)
            for header_row, column in header_cells:
                for row in worksheet.iter_rows(min_row=header_row + 1, min_col=column, max_col=column):
                    value = row[0].value
                    if _looks_like_manual_identifier(value):
                        identifiers.add(clean_text(value))
    finally:
        workbook.close()
    return identifiers, matched_header_count


def _count_matches(payload: bytes, identifiers: set[str]) -> int:
    if not identifiers:
        return 0
    text = payload.decode("utf-8", errors="ignore")
    return sum(text.count(identifier) for identifier in identifiers)


def _scan_xlsx(path: Path, identifiers: set[str]) -> int:
    total = 0
    try:
        with zipfile.ZipFile(path) as package:
            for member in package.infolist():
                if not member.is_dir():
                    total += _count_matches(package.read(member), identifiers)
    except (OSError, zipfile.BadZipFile):
        return _count_matches(path.read_bytes(), identifiers)
    return total


def _scan_output_directory(output_directory: Path, identifiers: set[str], json_out: Path) -> tuple[list[ScanFinding], int]:
    findings: list[ScanFinding] = []
    file_count = 0
    resolved_report = json_out.resolve()
    for path in sorted(output_directory.rglob("*")):
        if not path.is_file() or path.resolve() == resolved_report:
            continue
        file_count += 1
        relative_path = path.relative_to(output_directory)
        path_matches = sum(
            identifier in path_part
            for identifier in identifiers
            for path_part in relative_path.parts
        )
        content_matches = _scan_xlsx(path, identifiers) if path.suffix.lower() == ".xlsx" else _count_matches(path.read_bytes(), identifiers)
        artifact = "<redacted-path>" if path_matches else relative_path.as_posix()
        if path_matches:
            findings.append(ScanFinding("<redacted-path>", "path", path_matches))
        if content_matches:
            findings.append(
                ScanFinding(artifact, "content", content_matches)
            )
    return findings, file_count


def scan_legacy_identifiers(
    source_path: str | Path,
    output_directory: str | Path,
    identifier_headers: tuple[str, ...],
    json_out: str | Path,
) -> dict[str, Any]:
    source = Path(source_path)
    output = Path(output_directory)
    report_path = Path(json_out)
    if not source.is_file():
        raise FileNotFoundError(f"source workbook does not exist: {source}")
    if not output.is_dir():
        raise NotADirectoryError(f"output directory does not exist: {output}")
    identifiers, header_count = _extract_legacy_identifiers(source, identifier_headers)
    if not header_count:
        return {
            "schema_version": SCAN_SCHEMA_VERSION,
            "ok": False,
            "error_code": "SOURCE_IDENTIFIER_HEADER_MISSING",
            "source": {"sha256": sha256_file(source)},
            "identifier_header_count": 0,
            "legacy_identifier_count": 0,
            "files_scanned": 0,
            "findings": [],
        }
    findings, file_count = _scan_output_directory(output, identifiers, report_path)
    return {
        "schema_version": SCAN_SCHEMA_VERSION,
        "ok": not findings,
        "source": {"sha256": sha256_file(source)},
        "identifier_header_count": header_count,
        "legacy_identifier_count": len(identifiers),
        "files_scanned": file_count,
        "findings": [finding.to_dict() for finding in findings],
    }


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scan new outputs for legacy manual identifiers without echoing them.")
    parser.add_argument("source_xlsx")
    parser.add_argument("output_directory")
    parser.add_argument("--identifier-header", action="append", default=[])
    parser.add_argument("--json-out", required=True)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    try:
        args = _parse_args(argv)
        headers = tuple(args.identifier_header) or DEFAULT_IDENTIFIER_HEADERS
        payload = scan_legacy_identifiers(args.source_xlsx, args.output_directory, headers, args.json_out)
        write_json_no_overwrite(payload, args.json_out)
        print(json.dumps(payload, ensure_ascii=False, indent=2))
        if payload.get("error_code"):
            return 1
        return 0 if payload["ok"] else 2
    except (OSError, ValueError, zipfile.BadZipFile) as error:
        print(f"legacy identifier scan failed: {error}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
