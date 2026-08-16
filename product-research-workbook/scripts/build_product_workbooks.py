from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from validate_product_workbooks import ValidationReport, validate_contract, validate_process_workbook
from workbook_contract import (
    IDENTITY_FIELDS,
    PARAMETER_TRAILING_FIELDS,
    PROCUREMENT_BUSINESS_FIELDS,
    PROCUREMENT_FORMAL_HEADERS,
    PROCUREMENT_FORMAL_SHEET,
    REQUIRED_HYPERLINK_FIELDS,
    add_external_hyperlink,
    add_procurement_table,
    canonical_scalar,
    clean_text,
    coerce_formal_value,
    create_process_workbook,
    ensure_supported_openpyxl,
    is_blank,
    publish_file_no_overwrite,
    save_workbook_no_overwrite,
    sha256_file,
    style_formal_body,
    style_formal_parameter_sheet,
    style_formal_procurement_sheet,
)


def _print(payload: dict[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2))


def _validation_failure(report: ValidationReport) -> int:
    _print(report.to_dict())
    return 2


def _process_context(process_path: str | Path):
    report = validate_contract(process_path)
    if not report.ok:
        return None, report
    context_report = ValidationReport(mode="process", inputs=report.inputs)
    context = validate_process_workbook(process_path, context_report)
    return context, context_report


def _build_formal_workbook(process_path: str | Path) -> Workbook:
    context, context_report = _process_context(process_path)
    if context is None or not context_report.ok:
        raise ValueError("process workbook does not satisfy the confirmation gate")

    included_candidates = sorted(
        (
            candidate
            for candidate in context.candidates
            if candidate.inclusion == "已确认纳入" and candidate.product_id is not None
        ),
        key=lambda candidate: candidate.product_id or 0,
    )
    if not included_candidates:
        raise ValueError("no confirmed products are available to build a formal workbook")

    workbook = Workbook()
    workbook.remove(workbook.active)
    formal_products: dict[int, dict[str, Any]] = {}

    for sheet_name, fields in sorted(
        context.confirmed_fields_by_sheet.items(), key=lambda item: item[1][0].sheet_order
    ):
        worksheet = workbook.create_sheet(sheet_name)
        for column, field in enumerate(fields, start=1):
            worksheet.cell(1, column, field.field_name)
            worksheet.cell(2, column, field.unit or None)

        candidates = [
            candidate for candidate in included_candidates if candidate.target_sheet == sheet_name
        ]
        for candidate in candidates:
            row_number = worksheet.max_row + 1
            product_values: dict[str, Any] = {}
            for column, field in enumerate(fields, start=1):
                cell = worksheet.cell(row_number, column)
                if field.field_name == "产品编号":
                    value = candidate.product_id
                else:
                    evidence = context.main_evidence.get((candidate.candidate_id, field.field_name))
                    value = (
                        coerce_formal_value(evidence.formal_value, field.value_type)
                        if evidence is not None
                        else None
                    )
                    if evidence is not None and field.field_name in REQUIRED_HYPERLINK_FIELDS:
                        add_external_hyperlink(cell, evidence.source_url)
                cell.value = value
                product_values[field.field_name] = value
            formal_products[candidate.product_id] = product_values

        style_formal_parameter_sheet(worksheet, len(fields))
        style_formal_body(worksheet, 3)

    procurement = workbook.create_sheet(PROCUREMENT_FORMAL_SHEET)
    for column, header in enumerate(PROCUREMENT_FORMAL_HEADERS, start=1):
        procurement.cell(1, column, header)

    adopted_by_candidate: dict[str, list[Any]] = defaultdict(list)
    for record in context.adopted_procurement:
        adopted_by_candidate[record.candidate_id].append(record)
    for records in adopted_by_candidate.values():
        records.sort(key=lambda record: (record.channel_id, record.row))

    for candidate in included_candidates:
        product = formal_products[candidate.product_id]
        for record in adopted_by_candidate[candidate.candidate_id]:
            row_number = procurement.max_row + 1
            values = [
                candidate.product_id,
                product["产品名称"],
                product["产品型号"],
                *(record.business_values[field] for field in PROCUREMENT_BUSINESS_FIELDS),
            ]
            for column, value in enumerate(values, start=1):
                procurement.cell(row_number, column, value)
            for field_name in ("询价或购买链接", "国内代理/销售链接", "资料来源"):
                if field_name not in PROCUREMENT_BUSINESS_FIELDS:
                    continue
                cell = procurement.cell(
                    row_number, PROCUREMENT_FORMAL_HEADERS.index(field_name) + 1
                )
                if not is_blank(cell.value) and (
                    field_name in REQUIRED_HYPERLINK_FIELDS or str(cell.value).startswith(("http://", "https://"))
                ):
                    add_external_hyperlink(cell, clean_text(cell.value))

    add_procurement_table(procurement)
    style_formal_procurement_sheet(procurement)
    style_formal_body(procurement, 2)
    return workbook


def init_process(output: str | Path) -> int:
    ensure_supported_openpyxl()
    workbook = create_process_workbook()
    try:
        output_hash = save_workbook_no_overwrite(workbook, output)
    finally:
        workbook.close()
    _print(
        {
            "ok": True,
            "action": "init-process",
            "output": {"path": str(Path(output).resolve()), "sha256": output_hash},
        }
    )
    return 0


def build_formal(
    process: str | Path, output: str | Path, previous_formal: str | Path | None
) -> int:
    ensure_supported_openpyxl()
    output_path = Path(output)
    if output_path.exists():
        raise FileExistsError(f"output already exists: {output_path}")
    if not output_path.parent.is_dir():
        raise FileNotFoundError(f"output directory does not exist: {output_path.parent}")

    process_report = validate_contract(process)
    if not process_report.ok:
        return _validation_failure(process_report)

    workbook = _build_formal_workbook(process)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{output_path.stem}-", suffix=".xlsx", dir=output_path.parent
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        pre_publish_report = validate_contract(process, temporary_path, previous_formal)
        if not pre_publish_report.ok:
            return _validation_failure(pre_publish_report)
        output_hash = publish_file_no_overwrite(temporary_path, output_path)
        # Re-read the published path so the success report describes the file the
        # caller will actually consume, rather than the private staging name.
        final_report = validate_contract(process, output_path, previous_formal)
        if not final_report.ok:
            return _validation_failure(final_report)
    finally:
        workbook.close()
        temporary_path.unlink(missing_ok=True)

    _print(
        {
            "ok": True,
            "action": "build-formal",
            "output": {"path": str(output_path.resolve()), "sha256": output_hash},
            "validation": final_report.to_dict(),
        }
    )
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Create process workbooks and build formal product research workbooks."
    )
    subcommands = parser.add_subparsers(dest="command", required=True)
    init_parser = subcommands.add_parser("init-process")
    init_parser.add_argument("output", help="new process workbook path (.xlsx)")
    formal_parser = subcommands.add_parser("build-formal")
    formal_parser.add_argument("process", help="validated process workbook (.xlsx)")
    formal_parser.add_argument("output", help="new formal workbook path (.xlsx)")
    formal_parser.add_argument("--previous-formal", help="previous formal workbook for version checks")
    try:
        arguments = parser.parse_args(argv)
    except SystemExit as error:
        return 0 if error.code == 0 else 1

    try:
        if arguments.command == "init-process":
            return init_process(arguments.output)
        return build_formal(arguments.process, arguments.output, arguments.previous_formal)
    except (OSError, ValueError, RuntimeError) as error:
        _print({"ok": False, "code": "IO_OR_ARGUMENT_ERROR", "message": str(error)})
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
