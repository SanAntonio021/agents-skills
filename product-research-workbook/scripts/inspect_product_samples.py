from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from workbook_contract import canonical_scalar, clean_text, ensure_supported_openpyxl, sha256_file, write_json_no_overwrite


INVENTORY_SCHEMA_VERSION = "1.0"
DEFAULT_IDENTIFIER_HEADERS = ("产品编号",)


def _identifier_columns(worksheet: Any, headers: set[str]) -> list[tuple[int, int, str]]:
    columns: list[tuple[int, int, str]] = []
    for row in worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row)):
        for cell in row:
            header = clean_text(cell.value)
            if header in headers:
                columns.append((cell.row, cell.column, header))
    return columns


def _section_inventory(
    worksheet: Any,
    header_row: int,
    identifier_columns: set[int],
    sample_rows: int,
) -> dict[str, Any]:
    headers: list[tuple[int, str]] = []
    for column in range(1, worksheet.max_column + 1):
        if column in identifier_columns:
            continue
        header = clean_text(worksheet.cell(header_row, column).value)
        if header:
            headers.append((column, header))

    samples: list[dict[str, Any]] = []
    for row_number in range(header_row + 1, worksheet.max_row + 1):
        values = {
            header: canonical_scalar(worksheet.cell(row_number, column).value)
            for column, header in headers
            if worksheet.cell(row_number, column).value is not None
        }
        if not values:
            continue
        samples.append({"source_row": row_number, "values": values})
        if len(samples) >= sample_rows:
            break

    return {
        "header_row": header_row,
        "headers": [header for _, header in headers],
        "sample_rows": samples,
    }


def inspect_product_samples(
    source_path: str | Path,
    output_path: str | Path,
    identifier_headers: tuple[str, ...],
    sample_rows: int,
) -> dict[str, Any]:
    source = Path(source_path)
    if not source.is_file():
        raise FileNotFoundError(f"source workbook does not exist: {source}")
    if sample_rows < 1:
        raise ValueError("sample row count must be positive")

    wanted_headers = {header.strip() for header in identifier_headers if header.strip()}
    if not wanted_headers:
        raise ValueError("at least one identifier header is required")

    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    worksheets: list[dict[str, Any]] = []
    identifier_column_count = 0
    try:
        for worksheet in workbook.worksheets:
            found = _identifier_columns(worksheet, wanted_headers)
            identifier_column_count += len(found)
            sections: list[dict[str, Any]] = []
            for header_row in sorted({row for row, _, _ in found}):
                section_identifiers = [
                    {"column": column, "header": header}
                    for row, column, header in found
                    if row == header_row
                ]
                sections.append(
                    {
                        "identifier_columns": section_identifiers,
                        **_section_inventory(
                            worksheet,
                            header_row,
                            {item["column"] for item in section_identifiers},
                            sample_rows,
                        ),
                    }
                )
            worksheets.append(
                {
                    "worksheet": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "sections": sections,
                }
            )
    finally:
        workbook.close()

    if identifier_column_count == 0:
        raise ValueError("no configured identifier header was found in the source workbook")

    payload = {
        "schema_version": INVENTORY_SCHEMA_VERSION,
        "ok": True,
        "source": {"sha256": sha256_file(source)},
        "identifier_headers": sorted(wanted_headers),
        "identifier_column_count": identifier_column_count,
        "worksheets": worksheets,
    }
    output_hash = write_json_no_overwrite(payload, output_path)
    return {
        "ok": True,
        "action": "inspect-product-samples",
        "source": payload["source"],
        "identifier_column_count": identifier_column_count,
        "output": {"path": str(Path(output_path).resolve()), "sha256": output_hash},
    }


def _parse_args(argv: list[str] | None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Create a sample-first inventory without reading source identifier values into output."
    )
    parser.add_argument("source_xlsx")
    parser.add_argument("output_json")
    parser.add_argument("--identifier-header", action="append", default=[])
    parser.add_argument("--sample-rows", type=int, default=8)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    try:
        ensure_supported_openpyxl()
        args = _parse_args(argv)
        result = inspect_product_samples(
            args.source_xlsx,
            args.output_json,
            tuple(args.identifier_header) or DEFAULT_IDENTIFIER_HEADERS,
            args.sample_rows,
        )
        print(json.dumps(result, ensure_ascii=False, indent=2))
        return 0
    except (OSError, RuntimeError, ValueError) as error:
        print(
            json.dumps(
                {"ok": False, "code": "IO_OR_ARGUMENT_ERROR", "message": str(error)},
                ensure_ascii=False,
            )
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
